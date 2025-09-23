from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_from_directory, jsonify, send_file, abort
from flask_login import login_required, current_user
from io import BytesIO
from datetime import datetime, timedelta

from models import db, Plantel, Personal, Usuario, HistorialCambios, Delegacion, ObservacionPersonal as Observacion

from utils import registrar_historial, registrar_notificacion
from sqlalchemy import distinct, func, text
import pandas as pd
import re
from sqlalchemy.exc import IntegrityError
from authz import roles_required, requires, limit_query_to_user_delegacion, is_global_viewer, require_same_delegacion

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepInFrame, BaseDocTemplate, Frame, PageTemplate, NextPageTemplate, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm


personal_bp = Blueprint('personal_bp', __name__)

def _nombre_usuario_por_id(uid):
    if not uid:
        return "Usuario"
    u = Usuario.query.get(uid)
    return u.nombre if u else "Usuario"


def _check_access_persona(persona: Personal):
    if is_global_viewer():
        return
    # 1) intento directo
    delega_id = getattr(getattr(persona, "plantel", None), "delegacion_id", None)
    if delega_id is None:
        # 2) fallback por CCT -> Plantel
        pl = Plantel.query.filter_by(cct=(persona.cct or "").strip().upper()).first()
        delega_id = getattr(pl, "delegacion_id", None)
        if pl is None:
            abort(404)
    if delega_id != getattr(current_user, "delegacion_id", None):
        abort(403)

def _fetch_ficha_persona(persona_id: int):
    persona = Personal.query.get_or_404(persona_id)
    _check_access_persona(persona)

    plantel = persona.plantel
    delega  = plantel.delegacion if plantel else None

    # Observaciones (más recientes primero)
    obs = sorted(persona.observaciones, key=lambda o: o.fecha, reverse=True)

    # Historial (más recientes primero)
    hist = (HistorialCambios.query
            .filter_by(entidad='personal', entidad_id=persona.id)
            .order_by(HistorialCambios.fecha.desc())
            .all())

    datos = {
        "generado_por": getattr(current_user, "nombre", "Sistema"),
        "generado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "delegacion": getattr(delega, "nombre", ""),
        "nivel": getattr(delega, "nivel", ""),
        "plantel": {
            "cct": plantel.cct if plantel else (persona.cct or ""),
            "nombre": plantel.nombre if plantel else "",
            "direccion": f"{plantel.calle or ''} {plantel.num_exterior or ''}{(' Int. ' + plantel.num_interior) if plantel and plantel.num_interior else ''}, {plantel.colonia or ''}, {plantel.municipio or ''}, CP {plantel.cp or ''}" if plantel else ""
        },
        "persona": persona,
        "observaciones": obs,
        "historial": hist,
    }
    return datos

def _check_access_cct(cct: str):
    cct = (cct or "").strip().upper()  # ← normaliza
    plantel = Plantel.query.filter_by(cct=cct).first()
    if not plantel:
        abort(404)
    if not is_global_viewer():
        if plantel.delegacion_id != getattr(current_user, "delegacion_id", None):
            abort(403)
    return plantel

def _fetch_personal_detalle_por_cct(cct: str):
    """
    Devuelve:
    {
      'generado_por': str, 'generado_en': str,
      'delegacion': 'Nombre delegación', 'nivel': 'PRIMARIA ...',
      'plantel': {'cct':..., 'nombre':..., 'direccion':...},
      'rows': [ {todos los campos...}, ... ],
      'estadistica': {'H': int, 'M': int, 'T': int, 'funciones': {funcion: cnt, ...}}
    }
    """
    plantel = _check_access_cct(cct)
    delega = Delegacion.query.get(plantel.delegacion_id)

    # Trae personal del CCT
    personas = (Personal.query
        .filter(func.upper(Personal.cct) == cct.upper())
        .order_by(Personal.apellido_paterno.asc(), Personal.apellido_materno.asc(), Personal.nombre.asc())
        .all())

    rows = []
    hombres = mujeres = total = 0
    funciones = {}

    for p in personas:
        genero = (p.genero or '').upper()
        if genero == 'H':
            hombres += 1
        elif genero == 'M':
            mujeres += 1
        total += 1

        func = (p.funcion or 'SIN FUNCIÓN').upper()
        funciones[func] = funciones.get(func, 0) + 1

        rows.append({
            "apellido_paterno": p.apellido_paterno or "",
            "apellido_materno": p.apellido_materno or "",
            "nombre": p.nombre or "",
            "genero": p.genero or "",
            "rfc": p.rfc or "",
            "curp": p.curp or "",
            "clave_presupuestal": p.clave_presupuestal or "",
            "funcion": p.funcion or "",
            "grado_estudios": p.grado_estudios or "",
            "titulado": p.titulado or "",
            "fecha_ingreso": p.fecha_ingreso.strftime("%Y-%m-%d") if getattr(p, "fecha_ingreso", None) else "",
            "fecha_baja_jubilacion": p.fecha_baja_jubilacion.strftime("%Y-%m-%d") if getattr(p, "fecha_baja_jubilacion", None) else "",
            "estatus_membresia": p.estatus_membresia or "",
            "nombramiento": p.nombramiento or "",
            "domicilio": p.domicilio or "",
            "numero": p.numero or "",
            "localidad": p.localidad or "",
            "colonia": p.colonia or "",
            "municipio": p.municipio or "",
            "cp": p.cp or "",
            "tel1": p.tel1 or "",
            "tel2": p.tel2 or "",
            "correo_electronico": p.correo_electronico or "",
        })

    data = {
        "generado_por": getattr(current_user, "nombre", "Sistema"),
        "generado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "delegacion": delega.nombre if delega else "",
        "nivel": delega.nivel if delega else "",
        "plantel": {
            "cct": plantel.cct,
            "nombre": plantel.nombre or "",
            "direccion": f"{plantel.calle or ''} {plantel.num_exterior or ''}{(' Int. ' + plantel.num_interior) if plantel.num_interior else ''}, {plantel.colonia or ''}, {plantel.municipio or ''}, CP {plantel.cp or ''}"
        },
        "rows": rows,
        "estadistica": {"H": hombres, "M": mujeres, "T": total, "funciones": dict(sorted(funciones.items()))}
    }
    return data


@personal_bp.route("/busqueda_personal", methods=["GET", "POST"])
@login_required
@requires("personal.view")
def busqueda_personal():
    resultados = []
    if request.method == 'POST':
        filtros = {
            'apellido_paterno': request.form.get('apellido_paterno', '').strip(),
            'apellido_materno': request.form.get('apellido_materno', '').strip(),
            'nombre':            request.form.get('nombre', '').strip(),
            'curp':              request.form.get('curp', '').strip(),
            'rfc':               request.form.get('rfc', '').strip(),
            'domicilio':         request.form.get('domicilio', '').strip(),
            'colonia':           request.form.get('colonia', '').strip(),
        }

        consulta = limit_query_to_user_delegacion(Personal.query, Personal)  # 👈 scope
        for campo, valor in filtros.items():
            if valor:
                consulta = consulta.filter(getattr(Personal, campo).ilike(f"%{valor}%"))

        resultados = consulta.all()

    return render_template("busqueda_personal.html", resultados=resultados)


@personal_bp.route("/detalle_personal/<int:id>")
@login_required
@requires("personal.view")
def vista_detalle_personal(id):
    from flask import current_app
    persona = Personal.query.get_or_404(id)
    _check_access_persona(persona)  # 👈 valida delegación para roles acotados

    try:
        usuarios = Usuario.query.all()
    except Exception:
        current_app.logger.exception("Error cargando usuarios para detalle_personal")
        usuarios = []
    usuarios_por_id = {u.id: u for u in usuarios}

    try:
        niveles_disponibles = [
            n[0] for n in db.session.query(distinct(Plantel.nivel)).order_by(Plantel.nivel).all()
        ]
    except Exception:
        current_app.logger.exception("Error consultando distinct(Plantel.nivel) en detalle_personal")
        niveles_disponibles = []

    return render_template(
        "detalle_personal.html",
        persona=persona,
        niveles_disponibles=niveles_disponibles,
        usuarios_por_id=usuarios_por_id,
        timedelta=timedelta
    )



@personal_bp.route("/eliminar_personal/<int:id>", methods=["POST"])
@roles_required('admin')
def eliminar_personal(id):
    persona = Personal.query.get_or_404(id)

    # Guarda todo lo necesario ANTES de borrar (evita expire_on_commit)
    cct = persona.cct
    delegacion = persona.plantel.delegacion.nombre if persona.plantel and persona.plantel.delegacion else ''
    nombre_completo = f"{(persona.apellido_paterno or '').strip()} {(persona.apellido_materno or '').strip()} {(persona.nombre or '').strip()}".strip()
    curp = (persona.curp or "SIN CURP").strip()

    try:
        # 1) Borra dependientes primero (observaciones)
        Observacion.query.filter_by(personal_id=id).delete(synchronize_session=False)

        # 2) Borra el personal
        db.session.delete(persona)
        db.session.commit()

    except IntegrityError as e:
        db.session.rollback()
        flash("No se pudo eliminar: hay registros relacionados que lo impiden.", "danger")
        return redirect(url_for('personal_bp.vista_detalle_personal', id=id))
    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar: {e}", "danger")
        return redirect(url_for('personal_bp.vista_detalle_personal', id=id))

    # Notificación (usa los datos preservados)
    registrar_notificacion(
        f"{current_user.nombre} eliminó a {nombre_completo} ({curp}) de {cct}",
        tipo="personal"
    )

    flash("El personal ha sido eliminado correctamente.", "success")
    return redirect(url_for('personal_bp.vista_personal', delegacion=delegacion, cct=cct))


@personal_bp.route("/editar_personal/<int:id>", methods=["POST"])
@requires("personal.edit")
@require_same_delegacion(lambda id: Personal.query.get_or_404(id))
def editar_personal(id):
    persona = Personal.query.get_or_404(id)

    # ... (validaciones de CURP/RFC)

    campos = {
        "cct",  # <-- incluirlo
        "apellido_paterno","apellido_materno","nombre","genero","rfc","curp",
        "clave_presupuestal","funcion","grado_estudios","titulado",
        "fecha_ingreso","fecha_baja_jubilacion","estatus_membresia","nombramiento",
        "domicilio","numero","localidad","colonia","municipio","cp","tel1","tel2","correo_electronico",
        "num","dp_num_int","dp_cruce1","dp_cruce2",
        "escuela_nombre","turno","nivel","subs_modalidad","zona_escolar","sector",
        "dom_esc_calle","dom_esc_num_ext","dom_esc_num_int","dom_esc_cruce1","dom_esc_cruce2",
        "dom_esc_localidad","dom_esc_colonia","dom_esc_mun_nom","dom_esc_cp","dom_esc_coordenadas_gps",
        "estado","seccion_snte","del_o_ct","org","coord_reg","fun_sin",
    }

    if getattr(current_user, "rol", "") != "admin":
        campos.discard("cct")

    def parse_date(s):
        if not s: return None
        from datetime import datetime
        try: return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError: return None

    cambios = 0

    for campo in campos:
        if campo not in request.form:
            continue  # no tocar lo que no viene en el form

        raw = request.form.get(campo)

        if campo == "curp":
            valor_nuevo = (raw or "").strip().upper()
        elif campo == "rfc":
            valor_nuevo = (raw or "").strip().upper()
        elif campo in ("fecha_ingreso", "fecha_baja_jubilacion"):
            valor_nuevo = parse_date(raw)
            if raw and valor_nuevo is None:
                flash("Formato de fecha inválido (AAAA-MM-DD).", "danger")
                return redirect(url_for("personal_bp.vista_detalle_personal", id=persona.id))
        elif campo == "num":
            s = (raw or "").strip()
            valor_nuevo = int(s) if s.isdigit() else None
        elif campo == "cct":
            # si viene vacío, conserva el actual; y normaliza a MAYÚSCULAS
            valor_nuevo = (raw or persona.cct or "").strip().upper()
            actual = (persona.cct or "").strip().upper()
            if valor_nuevo != actual:
                if not Plantel.query.filter_by(cct=valor_nuevo).first():
                    flash("El CCT indicado no existe en planteles.", "danger")
                    return redirect(url_for("personal_bp.vista_detalle_personal", id=persona.id))
        else:
            valor_nuevo = (raw or "").strip()

        if hasattr(persona, campo):
            anterior = getattr(persona, campo, None)
            if str(anterior) != str(valor_nuevo):
                registrar_historial("personal", campo, anterior, valor_nuevo,
                                    persona.id, current_user.nombre, "edición")
                setattr(persona, campo, valor_nuevo)
                cambios += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al guardar: {e}", "danger")
        return redirect(url_for("personal_bp.vista_detalle_personal", id=persona.id))

    registrar_notificacion(
        f"{current_user.nombre} actualizó datos de {persona.nombre} ({getattr(persona,'curp','SIN CURP')})",
        tipo="personal"
    )
    flash(f"✅ Cambios guardados. ({cambios} campo(s) actualizado(s))", "success")
    return redirect(url_for("personal_bp.vista_detalle_personal", id=persona.id))



@personal_bp.route('/subir_excel_personal/<cct>', methods=['POST'])
@roles_required('admin', 'coordinador')
def subir_excel_personal(cct):
    plantel = _check_access_cct(cct)

    file = request.files.get('archivo_excel')
    if not file or file.filename == '':
        flash('Sube un archivo .xlsx en el campo "archivo_excel".', 'danger')
        return redirect(url_for('personal_bp.vista_personal',
                                delegacion=plantel.delegacion.nombre, cct=cct))

    try:
        import unicodedata, re
        df = pd.read_excel(file, sheet_name=0, engine="openpyxl", dtype=str)

        def norm(h):
            h = "" if h is None else str(h).strip()
            h = "".join(c for c in unicodedata.normalize("NFD", h) if unicodedata.category(c) != "Mn")
            h = re.sub(r"[^A-Za-z0-9]+","_", h).strip("_").lower()
            return h

        original_cols = list(df.columns)
        df.columns = [norm(c) for c in df.columns]

        # --- Aceptar variantes de encabezados:
        ALIASES = {
            "fch_baj_jub": "fecha_baja_jubilacion",
            "fecha_baja_por_jubilacion": "fecha_baja_jubilacion",
            "grado_max_estudios": "grado_maximo_estudios",
            "dom_esc_coords_gps": "dom_esc_coordenadas_gps",
        }
        df.rename(columns=ALIASES, inplace=True)

        # Debug útil para ver qué llegó
        flash("Encabezados normalizados: " + ", ".join(list(df.columns)[:50]) + ("..." if len(df.columns) > 50 else ""), "info")

        # --- Mapeo Excel -> campos de tu tabla
        MAP = {
            # Identificación
            "num": "num",
            "paterno": "apellido_paterno",
            "materno": "apellido_materno",
            "nombre": "nombre",
            "genero": "genero",
            "rfc": "rfc",
            "curp": "curp",
            "clave_presupuestal": "clave_presupuestal",
            "funcion": "funcion",
            "grado_maximo_estudios": "grado_estudios",
            "titulado": "titulado",
            "fecha_ingreso": "fecha_ingreso",
            "fecha_baja_jubilacion": "fecha_baja_jubilacion",
            "status_memb": "estatus_membresia",
            "nombramiento": "nombramiento",

            # Dirección persona
            "dp_calle": "domicilio",
            "dp_num_ext": "numero",
            "dp_num_int": "dp_num_int",
            "dp_cruce1": "dp_cruce1",
            "dp_cruce2": "dp_cruce2",
            "dp_localidad": "localidad",
            "dp_colonia": "colonia",
            "dp_mun_nom": "municipio",
            "dp_cp": "cp",
            "dp_tel1": "tel1",
            "dp_tel2": "tel2",
            "correo_electronico": "correo_electronico",

            # Escuela / sindical
            "escuela_nombre": "escuela_nombre",
            "cct": "cct",   # se sobreescribe al de la URL
            "turno": "turno",
            "nivel": "nivel",
            "subs_modalidad": "subs_modalidad",
            "zona_escolar": "zona_escolar",
            "sector": "sector",

            # Domicilio escuela
            "dom_esc_calle": "dom_esc_calle",
            "dom_esc_num_ext": "dom_esc_num_ext",
            "dom_esc_num_int": "dom_esc_num_int",
            "dom_esc_cruce1": "dom_esc_cruce1",
            "dom_esc_cruce2": "dom_esc_cruce2",
            "dom_esc_localidad": "dom_esc_localidad",
            "dom_esc_colonia": "dom_esc_colonia",
            "dom_esc_mun_nom": "dom_esc_mun_nom",
            "dom_esc_cp": "dom_esc_cp",
            "dom_esc_coordenadas_gps": "dom_esc_coordenadas_gps",

            # Otros
            "estado": "estado",
            "seccion_snte": "seccion_snte",
            "del_o_ct": "del_o_ct",
            "org": "org",
            "coord_reg": "coord_reg",
            "fun_sin": "fun_sin",
        }

        # Validación mínima (base)
        obligatorias = ["paterno","materno","nombre","genero","rfc","curp"]
        faltantes = [h for h in obligatorias if h not in df.columns]
        if faltantes:
            ejemplo = ", ".join(original_cols[:10])
            flash(f"Faltan columnas base: {', '.join(faltantes)}. Detectados (ejemplo): {ejemplo}", "danger")
            return redirect(url_for('personal_bp.vista_personal',
                                    delegacion=plantel.delegacion.nombre, cct=cct))

        # Aviso de columnas del MAP que no vinieron (para entender qué no se cargará)
        faltan_en_excel = [k for k in MAP.keys() if k not in df.columns]
        if faltan_en_excel:
            flash("Columnas esperadas no presentes (se ignorarán): " + ", ".join(faltan_en_excel), "warning")

        # Parseo de fechas robusto
        for k in ("fecha_ingreso", "fecha_baja_jubilacion"):
            if k in df.columns:
                df[k] = pd.to_datetime(df[k], errors="coerce").dt.date

        ok = bad = 0
        errores = []

        for i, row in df.iterrows():
            curp_val = (row.get("curp") or "").strip().upper()
            if not curp_val:
                bad += 1; errores.append(f"Fila {i+2}: CURP vacío.")
                continue

            dest = {}
            for src_norm, db_field in MAP.items():
                if src_norm not in df.columns:
                    continue
                val = row[src_norm]
                if pd.isna(val):
                    val = None

                # Normalizaciones puntuales
                if db_field == "genero" and val:
                    g = str(val).strip().upper()
                    val = "H" if g in ("H","MASCULINO","HOMBRE") else ("M" if g in ("M","F","FEMENINO","MUJER") else None)

                if db_field in ("cp","numero","tel1","tel2","cct","clave_presupuestal"):
                    val = None if val is None else str(val).strip()

                if db_field == "num":
                    s = "" if val is None else str(val).strip()
                    val = int(s) if s.isdigit() else None

                if db_field in ("rfc","curp") and val:
                    val = str(val).strip().upper()

                dest[db_field] = val

            # Fuerza adscripción al CCT de la URL
            dest["cct"] = plantel.cct

            # FK CCT
            if not Plantel.query.filter_by(cct=dest["cct"]).first():
                bad += 1; errores.append(f"Fila {i+2}: CCT {dest['cct']} no existe en plantel.")
                continue

            # UPSERT por CURP + CLAVE (multi-plaza)
            clave_val = (dest.get("clave_presupuestal") or "").strip().upper()
            dest["clave_presupuestal"] = clave_val  # normaliza también en el destino

            q = Personal.query.filter_by(curp=curp_val)
            if clave_val:
                q = q.filter_by(clave_presupuestal=clave_val)

            p = q.first()
            if not p:
                # Crea nuevo registro por combinación curp+clave
                p = Personal(curp=curp_val, clave_presupuestal=clave_val)
                db.session.add(p)


            # Asignar campos existentes
            for k, v in dest.items():
                if hasattr(Personal, k):
                    setattr(p, k, v)

            try:
                db.session.commit()
                ok += 1
            except Exception as e:
                db.session.rollback()
                bad += 1
                errores.append(f"Fila {i+2}: {e}")

        msg = f"Importación v2: {ok} OK, {bad} con error."
        if errores:
            flash(msg + " " + " | ".join(errores[:5]), "warning")
        else:
            flash(msg, "success")

        registrar_notificacion(
            f"{current_user.nombre} importó {ok} personas (errores: {bad}) desde Excel v2 en {cct}",
            tipo="personal"
        )
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al procesar el archivo: {e}", "danger")

    return redirect(url_for('personal_bp.vista_personal',
                            delegacion=plantel.delegacion.nombre, cct=cct))



@personal_bp.route("/historial/<entidad>/<int:id>")
@login_required
def ver_historial(entidad, id):
    historial = HistorialCambios.query.filter_by(entidad=entidad, entidad_id=id).order_by(HistorialCambios.fecha.desc()).all()
    return render_template("ver_historial.html", historial=historial, tipo=entidad)

@personal_bp.route('/agregar_observacion/<int:personal_id>', methods=['POST'])
@requires("personal.edit")
@require_same_delegacion(lambda personal_id: Personal.query.get_or_404(personal_id))
def agregar_observacion(personal_id):
    texto = (request.form.get('texto') or '').strip()
    if not texto:
        flash("El texto no puede estar vacío.", "warning")
        return redirect(request.referrer or url_for('personal_bp.vista_detalle_personal', id=personal_id))

    nueva_obs = Observacion(personal_id=personal_id, usuario_id=current_user.id, texto=texto)
    db.session.add(nueva_obs)
    db.session.commit()

    persona = Personal.query.get(personal_id)
    if persona:
        nombre = f"{(persona.apellido_paterno or '').strip()} {(persona.apellido_materno or '').strip()} {(persona.nombre or '').strip()}".strip()
        curp = (persona.curp or 'SIN CURP').strip()
        msg = f"{current_user.nombre} agregó una observación a {nombre} ({curp})"
    else:
        msg = f"{current_user.nombre} agregó una observación a personal #{personal_id}"

    registrar_notificacion(msg, tipo="personal")

    flash("✅ Observación registrada correctamente.", "success")
    return redirect(request.referrer or url_for('personal_bp.vista_detalle_personal', id=personal_id))

@personal_bp.route('/editar_observacion/<int:obs_id>', methods=['POST'])
@roles_required('admin')
def editar_observacion(obs_id):
    # 1) Obtener el personal_id SIN cargar el objeto a la sesión
    pid = db.session.query(Observacion.personal_id).filter_by(id=obs_id).scalar()
    if pid is None:
        flash("Observación no encontrada.", "warning")
        return redirect(request.referrer or url_for('personal_bp.vista_detalle_personal', id=0))

    # 2) Validar texto
    texto = (request.form.get('nuevo_texto') or request.form.get('texto') or '').strip()
    if not texto:
        flash("El texto no puede estar vacío.", "warning")
        return redirect(request.referrer or url_for('personal_bp.vista_detalle_personal', id=pid))

    try:
        # 3) 🔒 UPDATE dirigido con SQL crudo: SOLO 'texto'
        db.session.execute(
            text("UPDATE observacion_personal SET texto = :t WHERE id = :i"),
            {"t": texto, "i": obs_id}
        )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Error al actualizar observación: {e}", "danger")
        return redirect(request.referrer or url_for('personal_bp.vista_detalle_personal', id=pid))

    # 4) Notificación con nombre + CURP
    persona = Personal.query.get(pid)
    if persona:
        nombre = f"{(persona.apellido_paterno or '').strip()} {(persona.apellido_materno or '').strip()} {(persona.nombre or '').strip()}".strip()
        curp = (persona.curp or 'SIN CURP').strip()
        msg = f"{current_user.nombre} editó una observación de {nombre} ({curp})"
    else:
        msg = f"{current_user.nombre} editó una observación de personal #{pid}"

    registrar_notificacion(msg, tipo="personal")
    flash("✅ Observación actualizada correctamente.", "success")
    return redirect(request.referrer or url_for('personal_bp.vista_detalle_personal', id=pid))


@personal_bp.route('/eliminar_observacion/<int:obs_id>', methods=['POST'])
@roles_required('admin')
def eliminar_observacion(obs_id):
    obs = Observacion.query.get_or_404(obs_id)
    pid = obs.personal_id

    # Guarda datos para la noti antes de borrar
    persona = Personal.query.get(pid)
    if persona:
        nombre = f"{(persona.apellido_paterno or '').strip()} {(persona.apellido_materno or '').strip()} {(persona.nombre or '').strip()}".strip()
        curp = (persona.curp or 'SIN CURP').strip()
        msg = f"{current_user.nombre} eliminó una observación de {nombre} ({curp})"
    else:
        msg = f"{current_user.nombre} eliminó una observación de personal #{pid}"

    db.session.delete(obs)
    db.session.commit()

    registrar_notificacion(msg, tipo="personal")

    flash("Observación eliminada correctamente.", "success")
    return redirect(request.referrer or url_for('personal_bp.vista_detalle_personal', id=pid))



@personal_bp.route('/cambiar_adscripcion/<int:id>', methods=['POST'])
@roles_required('admin')
def cambiar_adscripcion(id):
    persona = Personal.query.get_or_404(id)

    # Normalizamos entradas
    nuevo_cct_raw = (request.form.get("nuevo_cct") or "")
    motivo = (request.form.get("motivo") or "").strip()
    nuevo_cct = nuevo_cct_raw.strip().upper()  # ← forzamos mayúsculas

    # Validaciones básicas
    if not nuevo_cct:
        flash("⚠️ Debes seleccionar un CCT válido.", "warning")
        return redirect(url_for("personal_bp.vista_detalle_personal", id=id))

    if nuevo_cct == ((persona.cct or "").strip().upper()):
        flash("ℹ️ El personal ya está adscrito a ese CCT.", "info")
        return redirect(url_for("personal_bp.vista_detalle_personal", id=id))

    if not motivo:
        flash("⚠️ Debes escribir un motivo del cambio.", "warning")
        return redirect(url_for("personal_bp.vista_detalle_personal", id=id))

    # ✅ Validar existencia del CCT destino ANTES de tocar nada
    dest_plantel = Plantel.query.filter_by(cct=nuevo_cct).first()
    if not dest_plantel:
        flash("⚠️ El CCT destino no existe.", "danger")
        return redirect(url_for("personal_bp.vista_detalle_personal", id=id))

    # Datos previos para auditoría
    cct_anterior = persona.cct
    estado_anterior = persona.estatus_membresia or ""

    # HISTORIAL: cambio de CCT (antes → después)
    registrar_historial(
        entidad="personal",
        campo="cct",
        valor_anterior=cct_anterior,
        valor_nuevo=nuevo_cct,
        entidad_id=persona.id,
        usuario=current_user.nombre,
        tipo="cambio de adscripción"
    )

    # OBSERVACIÓN: motivo del cambio (usando nombre del plantel destino)
    observacion = Observacion(
        personal_id=persona.id,
        usuario_id=current_user.id,
        texto=f"Motivo de cambio de adscripción: {motivo} (ahora adscrito al CCT {nuevo_cct} – {dest_plantel.nombre})",
        fecha=datetime.now()
    )
    db.session.add(observacion)

    # Aplicar cambio y normalizar estatus
    persona.cct = nuevo_cct
    persona.estatus_membresia = "ACTIVO"

    # HISTORIAL: rastro del estatus
    registrar_historial(
        entidad="personal",
        campo="estatus_membresia",
        valor_anterior=estado_anterior,
        valor_nuevo="ACTIVO (tras cambio de adscripción)",
        entidad_id=persona.id,
        usuario=current_user.nombre,
        tipo="ADSCRIPCION_CAMBIO"
    )

    # Guardar todo
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al actualizar adscripción: {e}", "danger")
        return redirect(url_for("personal_bp.vista_detalle_personal", id=persona.id))

    # Notificación
    registrar_notificacion(
        f"{current_user.nombre} cambió la adscripción de {persona.apellido_paterno} {persona.apellido_materno} {persona.nombre} "
        f"({getattr(persona, 'curp', 'SIN CURP')}) de {cct_anterior} a {nuevo_cct}. "
        f"Estatus regresó a ACTIVO.",
        tipo="personal"
    )

    flash("✅ Adscripción actualizada, estado normalizado y observación registrada.", "success")
    return redirect(url_for("personal_bp.vista_detalle_personal", id=persona.id))


@personal_bp.route('/personal/<delegacion>/<cct>')
@login_required
@requires("personal.view")  # 👈 protege también por permiso
def vista_personal(delegacion, cct):
    # 1) Valida acceso al CCT según el rol (secretario/admin ven todo; roles acotados solo su delegación)
    plantel = _check_access_cct(cct)

    # 2) Usa la delegación de la URL si existe; si no, cae a la del plantel
    delegacion_obj = Delegacion.query.filter(func.lower(Delegacion.nombre) == delegacion.lower()).first()
    if not delegacion_obj:
        delegacion_obj = plantel.delegacion

    # 3) Lista el personal del CCT, aplicando scope por delegación para roles acotados
    q = limit_query_to_user_delegacion(Personal.query, Personal)
    personal = (
        q.filter(func.upper(Personal.cct) == cct.upper())
        .order_by(Personal.apellido_paterno, Personal.apellido_materno, Personal.nombre)
        .all()
    )

    return render_template("consulta_personal.html", plantel=plantel, personal=personal, delegacion=delegacion_obj)


@personal_bp.route('/agregar_personal/<cct>', methods=['POST'])
@requires("personal.create")  # 👈 solo quien tenga permiso de crear puede entrar (secretario queda fuera)
def agregar_personal_manual(cct):
    # 1) 🔒 Validar que el usuario tenga acceso a ese CCT (secretario/admin ven todo; roles acotados solo su delegación)
    plantel = _check_access_cct(cct)

    def validar_curp(curp):
        return re.fullmatch(r'^[A-Z]{4}\d{6}[HM][A-Z]{5}[0-9A-Z]\d$', curp)

    def validar_rfc(rfc):
        return re.fullmatch(r'^[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}$', rfc)

    curp = (request.form.get('curp', '') or '').strip().upper()
    rfc  = (request.form.get('rfc', '') or '').strip().upper()
    clave = (request.form.get('clave_presupuestal', '') or '').strip().upper()  # 👈 normaliza

    if not validar_curp(curp):
        flash("⚠️ CURP inválido. Verifica el formato (18 caracteres, en mayúsculas).", "danger")
        return redirect(url_for('personal_bp.vista_personal', delegacion=plantel.delegacion.nombre, cct=cct))

    if not validar_rfc(rfc):
        flash("⚠️ RFC inválido. Verifica el formato (13 caracteres, en mayúsculas).", "danger")
        return redirect(url_for('personal_bp.vista_personal', delegacion=plantel.delegacion.nombre, cct=cct))

    if not clave:
        flash("⚠️ La clave presupuestal es obligatoria.", "danger")
        return redirect(url_for('personal_bp.vista_personal', delegacion=plantel.delegacion.nombre, cct=cct))

    # 👇 Evita duplicado por CURP + CLAVE (multi-plaza correcto)
    existe_curp_clave = Personal.query.filter_by(curp=curp, clave_presupuestal=clave).first()
    if existe_curp_clave:
        flash(
            f"❌ Ya existe este CURP con la clave {clave} para: "
            f"{existe_curp_clave.nombre} {existe_curp_clave.apellido_paterno} {existe_curp_clave.apellido_materno} "
            f"en el CCT {existe_curp_clave.cct}.",
            "danger"
        )
        return redirect(url_for('personal_bp.vista_personal', delegacion=plantel.delegacion.nombre, cct=cct))

    # 👇 Permite RFC repetido solo si es la misma persona (misma CURP)
    existe_rfc_otro = Personal.query.filter(Personal.rfc == rfc, Personal.curp != curp).first()
    if existe_rfc_otro:
        flash(
            f"❌ El RFC ya está registrado para: "
            f"{existe_rfc_otro.nombre} {existe_rfc_otro.apellido_paterno} {existe_rfc_otro.apellido_materno} "
            f"en el CCT {existe_rfc_otro.cct}.",
            "danger"
        )
        return redirect(url_for('personal_bp.vista_personal', delegacion=plantel.delegacion.nombre, cct=cct))

    def convertir_fecha(campo):
        valor = request.form.get(campo)
        if valor:
            try:
                return datetime.strptime(valor, "%Y-%m-%d").date()
            except ValueError:
                flash(f"⚠️ Fecha inválida en campo: {campo}. Usa el selector de fecha.", "warning")
        return None

    nuevo = Personal(
        cct=cct,  # 👈 fuerza adscripción al CCT validado
        apellido_paterno=(request.form.get('apellido_paterno', '') or '').strip(),
        apellido_materno=(request.form.get('apellido_materno', '') or '').strip(),
        nombre=(request.form.get('nombre', '') or '').strip(),
        genero=(request.form.get('genero', '') or '').strip(),
        rfc=rfc,
        curp=curp,
        clave_presupuestal=clave,
        funcion=(request.form.get('funcion', '') or '').strip(),
        grado_estudios=(request.form.get('grado_estudios', '') or '').strip(),
        titulado=(request.form.get('titulado', '') or '').strip(),
        fecha_ingreso=convertir_fecha('fecha_ingreso'),
        fecha_baja_jubilacion=convertir_fecha('fecha_baja_jubilacion'),
        estatus_membresia=(request.form.get('estatus_membresia', '') or '').strip(),
        nombramiento=(request.form.get('nombramiento', '') or '').strip(),
        domicilio=(request.form.get('domicilio', '') or '').strip(),
        numero=(request.form.get('numero', '') or '').strip(),
        localidad=(request.form.get('localidad', '') or '').strip(),
        colonia=(request.form.get('colonia', '') or '').strip(),
        municipio=(request.form.get('municipio', '') or '').strip(),
        cp=(request.form.get('cp', '') or '').strip(),
        tel1=(request.form.get('tel1', '') or '').strip(),
        tel2=(request.form.get('tel2', '') or '').strip(),
        correo_electronico=(request.form.get('correo_electronico', '') or '').strip()
    )

    # Cachear datos del plantel en el registro (consistencia de UI/reportes)
    p = plantel
    if p:
        nuevo.escuela_nombre = p.nombre
        nuevo.turno = p.turno
        nuevo.nivel = p.nivel
        nuevo.subs_modalidad = getattr(p, "modalidad", None)
        nuevo.zona_escolar = p.zona_escolar
        nuevo.sector = p.sector

        nuevo.dom_esc_calle = p.calle
        nuevo.dom_esc_num_ext = p.num_exterior
        nuevo.dom_esc_num_int = p.num_interior
        nuevo.dom_esc_cruce1 = p.cruce_1
        nuevo.dom_esc_cruce2 = p.cruce_2
        nuevo.dom_esc_localidad = p.localidad
        nuevo.dom_esc_colonia = p.colonia
        nuevo.dom_esc_mun_nom = p.municipio
        nuevo.dom_esc_cp = p.cp
        nuevo.dom_esc_coordenadas_gps = p.coordenadas_gps
        nuevo.estado = p.estado

    try:
        db.session.add(nuevo)
        db.session.commit()
        registrar_notificacion(
            f"{current_user.nombre} dio de alta a {nuevo.nombre} ({getattr(nuevo, 'curp', 'SIN CURP')}) en {cct}",
            tipo="personal"
        )
        flash("✅ Personal agregado correctamente.", "success")
    except Exception as e:
        db.session.rollback()
        flash("❌ Error al guardar el personal: " + str(e), "danger")
    # siempre redirige de vuelta al listado del plantel
    return redirect(url_for('personal_bp.vista_personal', delegacion=plantel.delegacion.nombre, cct=cct))



@personal_bp.route('/ccts_por_nivel')
@login_required
@requires("personal.view")
def obtener_ccts_por_nivel():
    nivel = request.args.get('nivel')
    if not nivel:
        return jsonify([])

    q = limit_query_to_user_delegacion(Plantel.query, Plantel)  # 👈 scope
    planteles = (q.filter(func.upper(Plantel.nivel) == nivel.upper())
                   .order_by(Plantel.nombre).all())

    return jsonify([{"cct": p.cct, "nombre": p.nombre} for p in planteles])


@personal_bp.route('/solicitar_baja/<int:id>', methods=['POST'])
@roles_required('admin', 'coordinador', 'delegado')
@require_same_delegacion(lambda id: Personal.query.get_or_404(id))
def solicitar_baja(id):
    persona = Personal.query.get_or_404(id)
    motivo = (request.form.get('motivo_baja') or '').strip()

    if not motivo:
        flash('Debes indicar el motivo de la baja.', 'warning')
        return redirect(url_for('personal_bp.vista_detalle_personal', id=id))

    # 1) Historial
    registrar_historial(
        entidad='personal',
        campo='solicitud_baja',
        valor_anterior='',
        valor_nuevo=f"Motivo: {motivo}",
        entidad_id=persona.id,
        usuario=getattr(current_user, 'nombre', None),
        tipo='BAJA'
    )

    # 2) Notificación
    nombre_completo = f"{persona.apellido_paterno} {persona.apellido_materno} {persona.nombre}".strip()
    cct_texto = getattr(persona, 'cct', None) or getattr(persona, 'plantel', None).cct

    descripcion_noti = (
        f"Solicitud de baja de personal: {nombre_completo} "
        f"(CCT {cct_texto}). Motivo: {motivo}. "
        f"Solicitado por: {getattr(current_user,'nombre','Usuario')}"
    )

    registrar_notificacion(descripcion_noti, tipo='personal')




    # 3) Marcar estado visible en UI
    persona.estatus_membresia = 'BAJA EN PROCESO'
    db.session.commit()

    flash('Solicitud de baja enviada y registrada. La ficha quedó en “Baja en proceso”.', 'success')

    # Redirige al listado del plantel (ajusta si usas otra ruta)
    try:
        return redirect(url_for('personal_bp.vista_personal',
                                delegacion=persona.plantel.delegacion.nombre,
                                cct=cct_texto))
    except Exception:
        return redirect(url_for('personal_bp.vista_detalle_personal', id=persona.id))
    

@personal_bp.route('/rechazar_baja/<int:id>', methods=['POST'])
@roles_required('admin', 'coordinador')
@require_same_delegacion(lambda id: Personal.query.get_or_404(id))
def rechazar_baja(id):
    persona = Personal.query.get_or_404(id)

    # Permite rechazar solo si realmente está en proceso
    if (persona.estatus_membresia or '').upper() != 'BAJA EN PROCESO':
        flash('Solo puedes rechazar bajas que estén en proceso.', 'warning')
        return redirect(url_for('personal_bp.vista_detalle_personal', id=id))

    motivo = (request.form.get('motivo_rechazo') or '').strip()
    if not motivo:
        flash('Debes indicar el motivo del rechazo.', 'warning')
        return redirect(url_for('personal_bp.vista_detalle_personal', id=id))

    estado_anterior = persona.estatus_membresia or ''

    # 1) Historial
    registrar_historial(
        entidad='personal',
        campo='rechazo_baja',
        valor_anterior=estado_anterior,
        valor_nuevo=f"Rechazada. Motivo: {motivo}",
        entidad_id=persona.id,
        usuario=getattr(current_user, 'nombre', None),
        tipo='BAJA_RECHAZADA'
    )

    # 2) Notificación
    nombre_completo = f"{persona.apellido_paterno} {persona.apellido_materno} {persona.nombre}".strip()
    cct_texto = getattr(persona, 'cct', None) or getattr(persona, 'plantel', None).cct
    descripcion_noti = (
        f"Rechazo de solicitud de baja para {nombre_completo} (CCT {cct_texto}). "
        f"Motivo: {motivo}. Por: {getattr(current_user,'nombre','Usuario')}."
    )
    registrar_notificacion(descripcion_noti, tipo='personal')

    # 3) Revertir estatus -> la tarjeta vuelve a color normal
    persona.estatus_membresia = 'ACTIVO'
    db.session.commit()

    flash('Solicitud de baja rechazada. Estatus regresó a ACTIVO.', 'success')

    try:
        return redirect(url_for('personal_bp.vista_personal',
                                delegacion=persona.plantel.delegacion.nombre,
                                cct=cct_texto))
    except Exception:
        return redirect(url_for('personal_bp.vista_detalle_personal', id=persona.id))
    
@personal_bp.route("/personal/<string:cct>/reporte/excel")
@login_required
@requires("personal.view")
def reporte_personal_cct_excel(cct):
    data = _fetch_personal_detalle_por_cct(cct)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabecera
    ws.merge_cells("A1:F1")
    ws["A1"] = "REPORTE DETALLADO DE PERSONAL POR CCT"
    ws["A1"].font = Font(size=14, bold=True); ws["A1"].alignment = center
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Delegación: {data['delegacion']} — {data['nivel']}  |  Plantel: {data['plantel']['nombre']} ({data['plantel']['cct']})"
    ws["A2"].alignment = center
    ws.merge_cells("A3:F3")
    ws["A3"] = f"Generado por: {data['generado_por']}  |  Fecha: {data['generado_en']}"
    ws["A3"].alignment = center

    ws.append([])
    ws.append(["Hombres", "Mujeres", "Total"])
    for c in ws[5]:
        c.font = bold; c.fill = header_fill; c.alignment = center; c.border = border
    ws.append([data["estadistica"]["H"], data["estadistica"]["M"], data["estadistica"]["T"]])
    ws.append([])

    # Tabla de funciones
    ws.append(["Función", "Cantidad"])
    ws["A8"].font = bold; ws["B8"].font = bold
    ws["A8"].fill = header_fill; ws["B8"].fill = header_fill
    ws["A8"].alignment = center; ws["B8"].alignment = center
    ws["A8"].border = ws["B8"].border = border

    fila = 9
    for f, cnt in data["estadistica"]["funciones"].items():
        ws.cell(row=fila, column=1, value=f).border = border
        ws.cell(row=fila, column=2, value=cnt).border = border
        fila += 1

    # Hoja detalle
    ws_det = wb.create_sheet("Detalle")
    headers = [
        "Apellido paterno","Apellido materno","Nombre","Género","RFC","CURP",
        "Clave presupuestal","Función","Grado estudios","Titulado",
        "Fecha ingreso","Fecha baja/jub","Estatus membresía","Nombramiento",
        "Domicilio","Número","Localidad","Colonia","Municipio","CP",
        "Tel 1","Tel 2","Correo"
    ]
    ws_det.append(headers)
    for cell in ws_det[1]:
        cell.font = bold; cell.fill = header_fill; cell.alignment = center; cell.border = border

    for r in data["rows"]:
        ws_det.append([
            r["apellido_paterno"], r["apellido_materno"], r["nombre"], r["genero"], r["rfc"], r["curp"],
            r["clave_presupuestal"], r["funcion"], r["grado_estudios"], r["titulado"],
            r["fecha_ingreso"], r["fecha_baja_jubilacion"], r["estatus_membresia"], r["nombramiento"],
            r["domicilio"], r["numero"], r["localidad"], r["colonia"], r["municipio"], r["cp"],
            r["tel1"], r["tel2"], r["correo_electronico"]
        ])

    # Bordes y ancho aproximado
    for row in ws_det.iter_rows(min_row=1, max_row=ws_det.max_row, min_col=1, max_col=len(headers)):
        for c in row:
            c.border = border
    for col_idx in range(1, len(headers)+1):
        max_len = 0
        for rr in range(1, ws_det.max_row+1):
            v = ws_det.cell(row=rr, column=col_idx).value
            max_len = max(max_len, len(str(v)) if v else 0)
        ws_det.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)

    ws_det.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output); output.seek(0)
    filename = f"personal_{data['plantel']['cct']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@personal_bp.route("/personal/<string:cct>/reporte/pdf")
@login_required
@requires("personal.view")
def reporte_personal_cct_pdf(cct):
    data = _fetch_personal_detalle_por_cct(cct)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(letter),
        leftMargin=20, rightMargin=20, topMargin=22, bottomMargin=22
    )

    styles = getSampleStyleSheet()
    small = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10.2,
        wordWrap="CJK",
        spaceBefore=0,
        spaceAfter=0,
    )
    small_bold = ParagraphStyle("small_bold", parent=small, fontName="Helvetica-Bold")
    title = ParagraphStyle("title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22)

    def P(txt, bold=False):
        return Paragraph("" if txt is None else str(txt), small_bold if bold else small)

    story = []

    # --- Encabezado / resumen ---
    story.append(Paragraph("<b>REPORTE DETALLADO DE PERSONAL POR CCT</b>", title))
    story.append(P(f"Delegación: {data['delegacion']} — {data['nivel']}  |  "
                   f"Plantel: {data['plantel']['nombre']} ({data['plantel']['cct']})"))
    story.append(P(f"Generado por: {data['generado_por']}  |  Fecha: {data['generado_en']}"))
    story.append(Spacer(1, 8))

    # Totales
    tot = Table(
        [[P("Hombres", True), P("Mujeres", True), P("Total", True)],
         [P(data["estadistica"]["H"]), P(data["estadistica"]["M"]), P(data["estadistica"]["T"])]],
        colWidths=[2.7*cm, 2.7*cm, 2.7*cm],
        hAlign="LEFT",
    )
    tot.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#ECECEC")),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ]))
    story.append(tot)
    story.append(Spacer(1, 8))

    # --- Tarjeta por persona (varias filas) ---
    col_w = doc.width / 3.0  # tres columnas iguales para las filas “compactas”

    for r in data["rows"]:
        nombre = f"{r['apellido_paterno']} {r['apellido_materno']} {r['nombre']}".strip()

        # Faja con el nombre (1 sola celda, ancho completo)
        head = Table([[P(nombre, True)]], colWidths=[doc.width], hAlign="LEFT")
        head.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F0F0F0")),
            ("BOX", (0,0), (-1,-1), 0.5, colors.black),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))

        # Filas compactas (3 columnas por fila). Nada de tablas anidadas en paralelo.
        filas = [
            [P(f"Género: {r['genero']}"), P(f"RFC: {r['rfc']}"), P(f"CURP: {r['curp']}")],
            [P(f"Función: {r['funcion']}"), P(f"Grado: {r['grado_estudios']}"), P(f"Titulado: {r['titulado']}")],
            [P(f"Clave presup.: {r['clave_presupuestal']}"), P(f"Estatus: {r['estatus_membresia']}"), P(f"Nombramiento: {r['nombramiento']}")],
            [P(f"Domicilio: {r['domicilio']} {r['numero'] or ''}"), P(f"Colonia: {r['colonia']}"), P(f"Municipio: {r['municipio']}")],
            [P(f"CP: {r['cp']}"), P(f"Tel1: {r['tel1']}"), P(f"Tel2: {r['tel2']}")],
            [P(f"Correo: {r['correo_electronico']}"), P(""), P("")],
            [P(f"F. ingreso: {r['fecha_ingreso']}"), P(f"F. baja/jub: {r['fecha_baja_jubilacion']}"), P("")],
        ]

        card = Table(filas, colWidths=[col_w, col_w, col_w], hAlign="LEFT")
        card.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING", (0,0), (-1,-1), 4),
            ("RIGHTPADDING", (0,0), (-1,-1), 4),
            ("TOPPADDING", (0,0), (-1,-1), 2),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ]))

        story.append(KeepTogether([head, card]))
        story.append(Spacer(1, 10))

    # Construir PDF
    doc.build(story)
    pdf = buf.getvalue()
    buf.close()

    filename = f"personal_{data['plantel']['cct']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(BytesIO(pdf), as_attachment=True, download_name=filename, mimetype="application/pdf")

@personal_bp.route("/personal/<int:persona_id>/ficha/pdf")
@login_required
@requires("personal.view")
def ficha_persona_pdf(persona_id):
    d = _fetch_ficha_persona(persona_id)
    p = d["persona"]

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter),
        leftMargin=20, rightMargin=20, topMargin=22, bottomMargin=22
    )

    styles = getSampleStyleSheet()
    small = ParagraphStyle("small", parent=styles["Normal"], fontName="Helvetica", fontSize=8.5, leading=10.2, wordWrap="CJK")
    small_bold = ParagraphStyle("small_bold", parent=small, fontName="Helvetica-Bold")
    title = ParagraphStyle("title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22)

    def P(txt, bold=False):
        return Paragraph("" if txt is None else str(txt), small_bold if bold else small)

    story = []
    # Encabezado
    story.append(Paragraph("<b>FICHA INDIVIDUAL DE PERSONAL</b>", title))
    story.append(P(f"Delegación: {d['delegacion']} — {d['nivel']}  |  Plantel: {d['plantel']['nombre']} ({d['plantel']['cct']})"))
    story.append(P(f"Generado por: {d['generado_por']}  |  Fecha: {d['generado_en']}"))
    story.append(Spacer(1, 8))

    # Cabecera con nombre grande
    nombre = f"{p.apellido_paterno or ''} {p.apellido_materno or ''} {p.nombre or ''}".strip()
    head = Table([[P(nombre, True)]], colWidths=[doc.width], hAlign="LEFT")
    head.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F0F0F0")),
        ("BOX", (0,0), (-1,-1), 0.5, colors.black),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(head)
    story.append(Spacer(1, 6))

    # Info básica en filas (3 columnas)
    col_w = doc.width / 3.0
    filas = [
        [P(f"Género: {p.genero or ''}"), P(f"RFC: {p.rfc or ''}"), P(f"CURP: {p.curp or ''}")],
        [P(f"Clave presup.: {p.clave_presupuestal or ''}"), P(f"Estatus: {p.estatus_membresia or ''}"), P(f"Nombramiento: {p.nombramiento or ''}")],
        [P(f"Función: {p.funcion or ''}"), P(f"Grado: {p.grado_estudios or ''}"), P(f"Titulado: {p.titulado or ''}")],
        [P(f"Domicilio: {p.domicilio or ''} {p.numero or ''}"), P(f"Colonia: {p.colonia or ''}"), P(f"Municipio: {p.municipio or ''}")],
        [P(f"CP: {p.cp or ''}"), P(f"Tel1: {p.tel1 or ''}"), P(f"Tel2: {p.tel2 or ''}")],
        [P(f"Correo: {p.correo_electronico or ''}"), P(""), P("")],
        [P(f"F. ingreso: {p.fecha_ingreso or ''}"), P(f"F. baja/jub: {p.fecha_baja_jubilacion or ''}"), P("")],
    ]
    card = Table(filas, colWidths=[col_w, col_w, col_w], hAlign="LEFT")
    card.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    story.append(card)
    story.append(Spacer(1, 10))

    # Observaciones
    story.append(Paragraph("<b>Observaciones</b>", small_bold))
    if d["observaciones"]:
        obs_rows = [[P("Fecha", True), P("Usuario", True), P("Texto", True)]]
        for o in d["observaciones"]:
            fecha = o.fecha.strftime("%Y-%m-%d %H:%M")
            usuario = _nombre_usuario_por_id(getattr(o, "usuario_id", None))
            obs_rows.append([P(fecha), P(usuario), P(o.texto or "")])
        obs_tbl = Table(obs_rows, colWidths=[3.0*cm, 5.0*cm, doc.width - 8.0*cm], hAlign="LEFT")
        obs_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#ECECEC")),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
        ]))
        story.append(obs_tbl)
    else:
        story.append(P("Sin observaciones registradas."))
    story.append(Spacer(1, 10))

    # Historial de cambios
    story.append(Paragraph("<b>Historial de movimientos</b>", small_bold))
    if d["historial"]:
        hist_rows = [[P("Fecha", True), P("Campo", True), P("Antes", True), P("Después", True), P("Usuario", True), P("Tipo", True)]]
        for h in d["historial"]:
            hist_rows.append([
                P(h.fecha.strftime("%Y-%m-%d %H:%M")),
                P(h.campo or ""),
                P(h.valor_anterior or ""),
                P(h.valor_nuevo or ""),
                P(h.usuario or ""),
                P(h.tipo or "")
            ])
        # ancho de columnas equilibrado
        cw = [3.2*cm, 3.0*cm, (doc.width-3.2*cm-3.0*cm-3.0*cm-3.0*cm)/2, (doc.width-3.2*cm-3.0*cm-3.0*cm-3.0*cm)/2, 3.0*cm, 3.0*cm]
        hist_tbl = Table(hist_rows, colWidths=cw, hAlign="LEFT", repeatRows=1)
        hist_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#ECECEC")),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
        ]))
        story.append(hist_tbl)
    else:
        story.append(P("Sin movimientos en el historial."))

    doc.build(story)
    pdf = buf.getvalue(); buf.close()
    filename = f"ficha_{p.apellido_paterno or ''}_{p.apellido_materno or ''}_{p.nombre or ''}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(BytesIO(pdf), as_attachment=True, download_name=filename, mimetype="application/pdf")


@personal_bp.route("/personal/<int:persona_id>/ficha/excel")
@login_required
@requires("personal.view")
def ficha_persona_excel(persona_id):
    d = _fetch_ficha_persona(persona_id)
    p = d["persona"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_row(key, value):
        r = ws.max_row + 1
        ws.cell(r, 1, key).font = bold
        ws.cell(r, 1).fill = header_fill
        ws.cell(r, 2, value)
        ws.cell(r, 1).border = border
        ws.cell(r, 2).border = border

    ws.merge_cells("A1:B1"); ws["A1"] = "FICHA INDIVIDUAL DE PERSONAL"; ws["A1"].font = Font(size=14, bold=True); ws["A1"].alignment = center
    set_row("Delegación", f"{d['delegacion']} — {d['nivel']}")
    set_row("Plantel", f"{d['plantel']['nombre']} ({d['plantel']['cct']})")
    set_row("Generado por", d["generado_por"])
    set_row("Fecha", d["generado_en"])
    ws.append([])

    nombre = f"{p.apellido_paterno or ''} {p.apellido_materno or ''} {p.nombre or ''}".strip()
    set_row("Nombre", nombre)
    set_row("Género", p.genero or "")
    set_row("RFC", p.rfc or "")
    set_row("CURP", p.curp or "")
    set_row("Clave presupuestal", p.clave_presupuestal or "")
    set_row("Función", p.funcion or "")
    set_row("Grado", p.grado_estudios or "")
    set_row("Titulado", p.titulado or "")
    set_row("Estatus", p.estatus_membresia or "")
    set_row("Nombramiento", p.nombramiento or "")
    set_row("Domicilio", f"{p.domicilio or ''} {p.numero or ''}")
    set_row("Colonia", p.colonia or "")
    set_row("Municipio", p.municipio or "")
    set_row("CP", p.cp or "")
    set_row("Tel1", p.tel1 or "")
    set_row("Tel2", p.tel2 or "")
    set_row("Correo", p.correo_electronico or "")
    set_row("Fecha ingreso", p.fecha_ingreso or "")
    set_row("Fecha baja/jub", p.fecha_baja_jubilacion or "")
    ws.append([])

    # Observaciones
    r0 = ws.max_row + 1
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=2)
    ws.cell(r0,1,"Observaciones").font = Font(bold=True); ws.cell(r0,1).fill = header_fill
    if d["observaciones"]:
        ws.append(["Fecha", "Usuario", "Texto"])
        for cell in ws[ws.max_row]:
            cell.font = bold; cell.fill = header_fill; cell.border = border
        for o in d["observaciones"]:
            ws.append([
                o.fecha.strftime("%Y-%m-%d %H:%M"),
                _nombre_usuario_por_id(getattr(o, "usuario_id", None)),
                o.texto or ""
            ])
            for cell in ws[ws.max_row]:
                cell.border = border
    else:
        ws.append(["", "Sin observaciones registradas."])


    ws.append([])
    # Historial
    r0 = ws.max_row + 1
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=2)
    ws.cell(r0,1,"Historial de movimientos").font = Font(bold=True); ws.cell(r0,1).fill = header_fill
    if d["historial"]:
        ws.append(["Fecha", "Campo", "Antes", "Después", "Usuario", "Tipo"])
        for cell in ws[ws.max_row]:
            cell.font = bold; cell.fill = header_fill; cell.border = border
        for h in d["historial"]:
            ws.append([
                h.fecha.strftime("%Y-%m-%d %H:%M"),
                h.campo or "", h.valor_anterior or "", h.valor_nuevo or "",
                h.usuario or "", h.tipo or ""
            ])
            for cell in ws[ws.max_row]:
                cell.border = border
    else:
        ws.append(["", "Sin movimientos en el historial."])

    # Anchos
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80
    for col in ["C","D","E","F"]:
        ws.column_dimensions[col].width = 30

    out = BytesIO(); wb.save(out); out.seek(0)
    filename = f"ficha_{p.apellido_paterno or ''}_{p.apellido_materno or ''}_{p.nombre or ''}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(out, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@personal_bp.route("/api/personal")
@requires("personal.view")
def api_listar_personal():
    return {"ok": False, "error": "No implementado"}, 501

@personal_bp.route("/api/personal", methods=["POST"])
@requires("personal.create")
def api_crear_persona():
    data = request.get_json() or {}

    cct = (data.get("cct") or "").strip().upper()

    if not cct:
        return {"ok": False, "error": "CCT requerido"}, 400

    # 🔒 valida acceso al CCT (y existencia)
    _check_access_cct(cct)

    # normalizaciones útiles
    if "curp" in data: data["curp"] = (data["curp"] or "").strip().upper()
    if "rfc"  in data: data["rfc"]  = (data["rfc"]  or "").strip().upper()
    if "clave_presupuestal" in data:
        data["clave_presupuestal"] = (data["clave_presupuestal"] or "").strip().upper()

    existe = Personal.query.filter_by(
        curp=data.get("curp"), 
        clave_presupuestal=data.get("clave_presupuestal")
    ).first()
    if existe:
        return {"ok": False, "error": "Ya existe CURP+clave_presupuestal"}, 409


    p = Personal(**data)
    db.session.add(p)
    db.session.commit()
    return {"ok": True, "id": p.id}


