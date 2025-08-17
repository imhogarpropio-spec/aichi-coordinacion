from flask import Blueprint, render_template, send_file, abort
from flask_login import login_required, current_user
from io import BytesIO
from models import db, Delegacion, Plantel, Personal, Notificacion, Usuario
from datetime import datetime
from collections import defaultdict
import zipfile

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

dashboard_bp = Blueprint('dashboard_bp', __name__)

@dashboard_bp.route('/dashboard')
@login_required
def dashboard():
    total_notificaciones = 0
    if current_user.rol == 'admin':
        total_notificaciones = Notificacion.query.filter_by(leida=False).count()

    return render_template(
        'dashboard.html',
        nombre=current_user.nombre,
        total_notificaciones=total_notificaciones
    )

# ---------- Helper: ficha PDF en bytes ----------
def _pdf_ficha_persona_bytes(persona: Personal, delegacion_nombre: str, nivel: str, plantel_dict: dict):
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter),
        leftMargin=20, rightMargin=20, topMargin=22, bottomMargin=22
    )

    styles = getSampleStyleSheet()
    small = ParagraphStyle("small", parent=styles["Normal"],
                           fontName="Helvetica", fontSize=8.5, leading=10.2, wordWrap="CJK")
    small_bold = ParagraphStyle("small_bold", parent=small, fontName="Helvetica-Bold")
    title = ParagraphStyle("title", parent=styles["Title"],
                           fontName="Helvetica-Bold", fontSize=18, leading=22)

    def P(txt, bold=False):
        return Paragraph("" if txt is None else str(txt), small_bold if bold else small)

    story = []
    story.append(Paragraph("<b>FICHA INDIVIDUAL DE PERSONAL</b>", title))
    story.append(P(f"Delegación: {delegacion_nombre} — {nivel}  |  "
                   f"Plantel: {plantel_dict.get('nombre','')} ({plantel_dict.get('cct','')})"))
    story.append(P(f"Generado por: {getattr(current_user, 'nombre', 'Sistema')}  |  "
                   f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"))
    story.append(Spacer(1, 8))

    nombre = f"{persona.apellido_paterno or ''} {persona.apellido_materno or ''} {persona.nombre or ''}".strip()
    head = Table([[P(nombre, True)]], colWidths=[doc.width], hAlign="LEFT")
    head.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F0F0F0")),
        ("BOX", (0,0), (-1,-1), 0.5, colors.black),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(head)
    story.append(Spacer(1, 6))

    col_w = doc.width / 3.0
    filas = [
        [P(f"Género: {persona.genero or ''}"), P(f"RFC: {persona.rfc or ''}"), P(f"CURP: {persona.curp or ''}")],
        [P(f"Clave presup.: {persona.clave_presupuestal or ''}"), P(f"Estatus: {persona.estatus_membresia or ''}"), P(f"Nombramiento: {persona.nombramiento or ''}")],
        [P(f"Función: {persona.funcion or ''}"), P(f"Grado: {persona.grado_estudios or ''}"), P(f"Titulado: {persona.titulado or ''}")],
        [P(f"Domicilio: {persona.domicilio or ''} {persona.numero or ''}"), P(f"Colonia: {persona.colonia or ''}"), P(f"Municipio: {persona.municipio or ''}")],
        [P(f"CP: {persona.cp or ''}"), P(f"Tel1: {persona.tel1 or ''}"), P(f"Tel2: {persona.tel2 or ''}")],
        [P(f"Correo: {persona.correo_electronico or ''}"), P(""), P("")],
        [P(f"F. ingreso: {persona.fecha_ingreso or ''}"), P(f"F. baja/jub: {persona.fecha_baja_jubilacion or ''}"), P("")],
    ]
    card = Table(filas, colWidths=[col_w, col_w, col_w], hAlign="LEFT")
    card.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    story.append(KeepTogether([card]))

    doc.build(story)
    pdf = buf.getvalue(); buf.close()
    return pdf

# ---------- Helper: Excel general en bytes ----------
def _excel_reporte_general_bytes(personal_list, planteles_map, delegaciones_map):
    wb = Workbook()
    ws_res_del = wb.active; ws_res_del.title = "Resumen delegaciones"
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 1) Resumen por delegación
    ws_res_del.append(["Delegación","Nivel","Personal"])
    for c in ws_res_del[1]:
        c.font = bold; c.fill = header_fill; c.alignment = center; c.border = border

    conteo_del = defaultdict(int)
    for p in personal_list:
        plantel = planteles_map.get(p.cct)
        if plantel:
            del_id = plantel['delegacion_id']
            conteo_del[del_id] += 1

    for del_id, cnt in sorted(conteo_del.items(), key=lambda x: delegaciones_map.get(x[0],{}).get("nombre","")):
        d = delegaciones_map.get(del_id, {})
        ws_res_del.append([d.get("nombre",""), d.get("nivel",""), cnt])

    ws_res_del.column_dimensions["A"].width = 38
    ws_res_del.column_dimensions["B"].width = 20
    ws_res_del.column_dimensions["C"].width = 14

    # 2) Resumen por CCT
    ws_cct = wb.create_sheet("Resumen CCT")
    ws_cct.append(["CCT","Plantel","Delegación","Nivel","Total personal"])
    for c in ws_cct[1]:
        c.font = bold; c.fill = header_fill; c.alignment = center; c.border = border

    conteo_cct = defaultdict(int)
    for p in personal_list:
        conteo_cct[p.cct] += 1
    for cct, cnt in sorted(conteo_cct.items()):
        pl = planteles_map.get(cct, {})
        delg = delegaciones_map.get(pl.get("delegacion_id"), {})
        ws_cct.append([cct, pl.get("nombre",""), delg.get("nombre",""), delg.get("nivel",""), cnt])

    for i, w in enumerate([14, 38, 32, 18, 16], start=1):
        ws_cct.column_dimensions[get_column_letter(i)].width = w

    # 3) Detalle Personal
    ws_det = wb.create_sheet("Detalle personal")
    headers = [
        "CCT","Plantel","Delegación","Nivel",
        "Apellido paterno","Apellido materno","Nombre","Género","RFC","CURP",
        "Clave presupuestal","Función","Grado","Titulado",
        "Fecha ingreso","Fecha baja/jub","Estatus membresía","Nombramiento",
        "Domicilio","Número","Localidad","Colonia","Municipio","CP",
        "Tel1","Tel2","Correo"
    ]
    ws_det.append(headers)
    for c in ws_det[1]:
        c.font = bold; c.fill = header_fill; c.alignment = center; c.border = border

    for p in personal_list:
        pl = planteles_map.get(p.cct, {})
        delg = delegaciones_map.get(pl.get("delegacion_id"), {})
        ws_det.append([
            p.cct or "", pl.get("nombre",""), delg.get("nombre",""), delg.get("nivel",""),
            p.apellido_paterno or "", p.apellido_materno or "", p.nombre or "", p.genero or "", p.rfc or "", p.curp or "",
            p.clave_presupuestal or "", p.funcion or "", p.grado_estudios or "", p.titulado or "",
            getattr(p, "fecha_ingreso", None) or "", getattr(p, "fecha_baja_jubilacion", None) or "", p.estatus_membresia or "", p.nombramiento or "",
            p.domicilio or "", p.numero or "", p.localidad or "", p.colonia or "", p.municipio or "", p.cp or "",
            p.tel1 or "", p.tel2 or "", p.correo_electronico or ""
        ])

    for col_idx in range(1, len(headers)+1):
        ws_det.column_dimensions[get_column_letter(col_idx)].width = 18
    ws_det.freeze_panes = "A2"

    # 4) Funciones (conteo general)
    ws_fun = wb.create_sheet("Funciones")
    ws_fun.append(["Función","Cantidad"])
    for c in ws_fun[1]:
        c.font = bold; c.fill = header_fill; c.alignment = center; c.border = border

    conteo_fun = defaultdict(int)
    for p in personal_list:
        fn = (p.funcion or "SIN FUNCIÓN").upper()
        conteo_fun[fn] += 1
    for fn, cnt in sorted(conteo_fun.items()):
        ws_fun.append([fn, cnt])
    ws_fun.column_dimensions["A"].width = 40
    ws_fun.column_dimensions["B"].width = 14

    out = BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

@dashboard_bp.route('/dashboard/reporte_general.zip')
@login_required
def reporte_general_zip():
    if current_user.rol != 'admin':
        abort(403)

    # Carga base
    delegaciones = Delegacion.query.all()
    planteles = Plantel.query.all()
    personal = Personal.query.order_by(Personal.apellido_paterno,
                                       Personal.apellido_materno,
                                       Personal.nombre).all()

    # Mapas rápidos
    delegaciones_map = {d.id: {"nombre": d.nombre, "nivel": d.nivel} for d in delegaciones}
    planteles_map = {p.cct: {"cct": p.cct, "nombre": p.nombre, "delegacion_id": p.delegacion_id} for p in planteles}

    # 1) Excel general
    excel_bytes = _excel_reporte_general_bytes(personal, planteles_map, delegaciones_map)

    # 2) PDFs individuales por persona (en carpetas por CCT)
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        # Excel
        zf.writestr("reporte_general.xlsx", excel_bytes)

        # PDFs
        for p in personal:
            pl = planteles_map.get(p.cct, {})
            delg = delegaciones_map.get(pl.get("delegacion_id"), {})
            pdf_bytes = _pdf_ficha_persona_bytes(
                persona=p,
                delegacion_nombre=delg.get("nombre",""),
                nivel=delg.get("nivel",""),
                plantel_dict=pl or {}
            )
            safe_name = f"{(p.apellido_paterno or '').strip()}_{(p.apellido_materno or '').strip()}_{(p.nombre or '').strip()}".replace(' ','_')
            path = f"fichas_pdf/{p.cct or 'SIN_CCT'}/{safe_name}.pdf"
            zf.writestr(path, pdf_bytes)

    zip_buf.seek(0)
    filename = f"reporte_general_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return send_file(zip_buf, as_attachment=True, download_name=filename, mimetype="application/zip")