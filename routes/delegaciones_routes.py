from flask import Blueprint, render_template, request, redirect, url_for, flash, send_from_directory, send_file, abort, jsonify
from flask_login import login_required, current_user
from io import BytesIO
from datetime import datetime
from utils import registrar_notificacion, registrar_historial
from models import db, Delegacion, Plantel, Personal
import pandas as pd
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, not_
from authz import roles_required, has_role
from pytz import timezone
import json
from math import ceil



# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import PageBreak
from reportlab.lib.units import cm


delegaciones_bp = Blueprint('delegaciones_bp', __name__)

def _parse_tabulator_args(req):
    page = req.args.get("page", type=int) or 1
    # default 100 y límite entre 1 y 500
    size = req.args.get("size", type=int) or 100
    size = max(1, min(size, 500))

    # sorters[n][field], sorters[n][dir]
    sorters = []
    i = 0
    while True:
        field = req.args.get(f"sorter[{i}][field]")
        if field is None:
            break
        sorters.append({
            "field": field,
            "dir": req.args.get(f"sorter[{i}][dir]", "asc")
        })
        i += 1

    # filter[n][field], filter[n][value]
    filters = []
    i = 0
    while True:
        field = req.args.get(f"filter[{i}][field]")
        if field is None:
            break
        filters.append({
            "field": field,
            "value": req.args.get(f"filter[{i}][value]")
        })
        i += 1

    return page, size, sorters, filters


def _alcance_delegaciones_query():
    """Devuelve el query base de delegaciones respetando el rol."""
    if current_user.rol == 'delegado':
        return Delegacion.query.filter(Delegacion.id == current_user.delegacion_id)
    # admin/coordinador: opcional ?delegacion_id=...
    delegacion_id = request.args.get("delegacion_id", type=int)
    if delegacion_id:
        return Delegacion.query.filter(Delegacion.id == delegacion_id)
    return Delegacion.query

def _fetch_personal_por_cct():
    """
    Estructura:
    {
      'generado_por': str,
      'generado_en': str,
      'delegaciones': [
         {
           'id': int, 'nombre': str, 'nivel': str,
           'planteles': [
              {
                'cct': str, 'plantel': str,
                'hombres': int, 'mujeres': int, 'total': int,
                'funciones': { 'DOCENTE': 10, 'DIRECTOR': 1, ... }
              }, ...
           ],
           'totales': { 'hombres': int, 'mujeres': int, 'total': int, 'funciones': {...} },
           'funciones_orden': [ 'DOCENTE', 'DIRECTOR', ... ]  # orden normalizado
         }, ...
      ]
    }
    """
    del_q = _alcance_delegaciones_query().order_by(Delegacion.nivel.asc(), Delegacion.nombre.asc())

    resultado = {
        "generado_por": getattr(current_user, "nombre", "Sistema"),
        "generado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "delegaciones": []
    }

    for d in del_q.all():
        # Planteles de la delegación
        planteles = (Plantel.query
                     .filter(Plantel.delegacion_id == d.id)
                     .order_by(Plantel.nombre.asc())
                     .all())

        # Mapa CCT -> nombre de plantel
        cct2nombre = {p.cct: (p.nombre or "") for p in planteles if p.cct}

        # Personal por CCT (join por CCT)
        per_rows = (
            db.session.query(
                Personal.cct,
                Personal.genero,
                Personal.funcion_coordinacion,
                func.count(Personal.id)
            )
            .join(Plantel, Personal.cct == Plantel.cct)
            .filter(Plantel.delegacion_id == d.id)
            .group_by(Personal.cct, Personal.genero, Personal.funcion_coordinacion)
            .all()
        )

        # Acumular por CCT (usa funcion_coordinacion)
        por_cct = {}
        funciones_globales = set()
        for cct, genero, funcion_coord, cnt in per_rows:
            if not cct:
                continue
            nodo = por_cct.setdefault(cct, {
                "cct": cct,
                "plantel": cct2nombre.get(cct, ""),
                "hombres": 0, "mujeres": 0, "total": 0,
                "funciones": {}
            })
            g = (genero or "").upper()
            if g == "H":
                nodo["hombres"] += int(cnt)
            elif g == "M":
                nodo["mujeres"] += int(cnt)
            nodo["total"] += int(cnt)

            f = (funcion_coord or "SIN FUNCIÓN COORD.").upper()
            nodo["funciones"][f] = nodo["funciones"].get(f, 0) + int(cnt)
            funciones_globales.add(f)

        # Asegurar presencia de CCTs sin personal (con ceros)
        for cct, nombre in cct2nombre.items():
            por_cct.setdefault(cct, {
                "cct": cct, "plantel": nombre,
                "hombres": 0, "mujeres": 0, "total": 0,
                "funciones": {}
            })

        # Orden de funciones estable
        funciones_orden = sorted(funciones_globales)

        # Totales por delegación
        tot_h = sum(n["hombres"] for n in por_cct.values())
        tot_m = sum(n["mujeres"] for n in por_cct.values())
        tot_tot = sum(n["total"] for n in por_cct.values())
        tot_func = {}
        for n in por_cct.values():
            for f, v in n["funciones"].items():
                tot_func[f] = tot_func.get(f, 0) + v

        resultado["delegaciones"].append({
            "id": d.id,
            "nombre": d.nombre,
            "nivel": d.nivel,
            "planteles": sorted(por_cct.values(), key=lambda x: x["plantel"]),
            "totales": {"hombres": tot_h, "mujeres": tot_m, "total": tot_tot, "funciones": tot_func},
            "funciones_orden": funciones_orden
        })

    return resultado

def _fetch_ccts_grouped_by_delegacion():
    """
    Regresa:
      {
        "generado_por": str,
        "generado_en": str,
        "delegaciones": [
            {
              "id": int, "nombre": str, "nivel": str, "delegado": str|None,
              "planteles": [ {campos del Plantel...}, ... ],
            }, ...
        ],
        "totales": {
            "delegaciones": int,
            "planteles": int,
            "planteles_por_delegacion": {nombre_delegacion: cantidad}
        }
      }
    Alcance:
      - delegado: solo su delegación
      - admin/coordinador: todas o una en específico via ?delegacion_id=ID
    """
    # Alcance por rol
    if current_user.rol == 'delegado':
        delegaciones_q = Delegacion.query.filter(Delegacion.id == current_user.delegacion_id)
    else:
        delegacion_id = request.args.get("delegacion_id", type=int)
        if delegacion_id:
            delegaciones_q = Delegacion.query.filter(Delegacion.id == delegacion_id)
        else:
            delegaciones_q = Delegacion.query

    delegaciones_q = delegaciones_q.order_by(Delegacion.nivel.asc(), Delegacion.nombre.asc())
    delegaciones = []
    total_planteles = 0
    planteles_por_delegacion = {}

    for d in delegaciones_q.all():
        # obtén planteles ordenados por nombre (ajusta si prefieres por CCT)
        planteles = (Plantel.query
                     .filter(Plantel.delegacion_id == d.id)
                     .order_by(Plantel.nombre.asc())
                     .all())

        rows = []
        for p in planteles:
            rows.append({
                "cct": (p.cct or "").upper(),
                "nombre": p.nombre or "",
                "turno": p.turno or "",
                "nivel": p.nivel or "",
                "modalidad": p.modalidad or "",
                "zona_escolar": p.zona_escolar or "",
                "sector": p.sector or "",
                "calle": p.calle or "",
                "num_exterior": p.num_exterior or "",
                "num_interior": p.num_interior or "",
                "cruce_1": p.cruce_1 or "",
                "cruce_2": p.cruce_2 or "",
                "localidad": p.localidad or "",
                "colonia": p.colonia or "",
                "municipio": p.municipio or "",
                "cp": p.cp or "",
                "coordenadas_gps": p.coordenadas_gps or "",
            })

        delegaciones.append({
            "id": d.id,
            "nombre": d.nombre,
            "nivel": d.nivel,
            "delegado": d.delegado,
            "planteles": rows
        })
        planteles_por_delegacion[d.nombre] = len(rows)
        total_planteles += len(rows)

    data = {
        "generado_por": getattr(current_user, "nombre", "Sistema"),
        "generado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "delegaciones": delegaciones,
        "totales": {
            "delegaciones": len(delegaciones),
            "planteles": total_planteles,
            "planteles_por_delegacion": planteles_por_delegacion
        }
    }
    return data


def _fetch_delegaciones_data_para_reporte():
    """
    Regresa un dict con:
      - total_delegaciones
      - totales_por_nivel {nivel: cantidad}
      - rows: lista de tuplas (id, nombre, nivel, delegado, num_ccts)
      - generado_por, generado_en
    Respeta el rol 'delegado' (solo ve su delegación).
    """
    # Base query según rol
    if current_user.rol == 'delegado':
        base_q = Delegacion.query.filter(Delegacion.id == current_user.delegacion_id)
    else:
        base_q = Delegacion.query

    # Totales por nivel
    totales_por_nivel = dict(
        base_q.with_entities(Delegacion.nivel, func.count(Delegacion.id))
              .group_by(Delegacion.nivel)
              .all()
    )

    # Subquery: conteo de planteles por delegacion
    subq = (
        db.session.query(Plantel.delegacion_id, func.count(Plantel.id).label('num_ccts'))
        .group_by(Plantel.delegacion_id)
        .subquery()
    )

    rows = (
        base_q.with_entities(
            Delegacion.id,
            Delegacion.nombre,
            Delegacion.nivel,
            Delegacion.delegado,
            func.coalesce(subq.c.num_ccts, 0).label('num_ccts')
        )
        .outerjoin(subq, Delegacion.id == subq.c.delegacion_id)
        .order_by(Delegacion.nivel.asc(), Delegacion.nombre.asc())
        .all()
    )

    data = {
        "total_delegaciones": len(rows),
        "totales_por_nivel": totales_por_nivel,
        "rows": rows,
        "generado_por": getattr(current_user, "nombre", "Sistema"),
        "generado_en": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return data

@delegaciones_bp.route('/delegaciones', methods=['GET'])
@login_required
def vista_delegaciones():
    # Delegado: solo su propia delegación
    if current_user.rol == 'delegado':
        delegaciones = Delegacion.query.filter(
            Delegacion.id == current_user.delegacion_id
        ).all()
    else:
        delegaciones = Delegacion.query.order_by(Delegacion.nivel, Delegacion.nombre).all()

    # Agrupar por nivel (como ya lo tenías)
    niveles = {}
    for d in delegaciones:
        niveles.setdefault(d.nivel, []).append(d)

    niveles_disponibles = [
        'PREESCOLAR GENERAL', 'PRIMARIA GENERAL', 'SECUNDARIA GENERAL',
        'SECUNDARIAS TÉCNICAS', 'TELESECUNDARIAS', 'NIVELES ESPECIALES',
        'MEDIA SUPERIOR', 'JUBILADOS'
    ]
    return render_template('consulta_delegaciones.html',
                           niveles=niveles,
                           niveles_disponibles=niveles_disponibles)

@delegaciones_bp.route('/eliminar_delegacion/<int:id>', methods=['POST'])
@roles_required('admin', 'coordinador')
def eliminar_delegacion(id):
    deleg = Delegacion.query.get_or_404(id)
    nombre = deleg.nombre
    nivel = deleg.nivel

    try:
        db.session.delete(deleg)
        db.session.commit()

        # ✅ Notificación correcta (delegación, no cct)
        registrar_notificacion(
            f"{current_user.nombre} eliminó la delegación '{nombre}' (nivel {nivel})",
            tipo="delegacion"
        )

        flash(f"Delegación '{nombre}' eliminada.", "success")

    except IntegrityError:
        db.session.rollback()
        flash(f"No se puede eliminar '{nombre}' porque tiene planteles asociados.", "warning")
        # Si quieres permitir borrado en cascada, hay que configurar el modelo con cascade / ondelete.

    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar la delegación: {e}", "danger")

    return redirect(url_for('delegaciones_bp.vista_delegaciones'))


@delegaciones_bp.route('/editar_delegacion/<int:id>', methods=['POST'])
@roles_required('admin', 'coordinador')
def editar_delegacion(id):
    deleg = Delegacion.query.get_or_404(id)
    deleg.nombre = request.form['nuevo_nombre']
    deleg.nivel = request.form['nuevo_nivel']
    deleg.delegado = request.form['nuevo_delegado']
    db.session.commit()
    registrar_notificacion(
        f"{current_user.nombre} actualizó la delegación '{deleg.nombre}' (nivel {deleg.nivel})",
        tipo="delegacion"
    )
    flash("Delegación actualizada correctamente.", "success")
    return redirect(url_for('delegaciones_bp.vista_delegaciones'))

@delegaciones_bp.route('/delegacion/<int:delegacion_id>/ccts', methods=['GET', 'POST'])
@login_required
def vista_ccts_por_delegacion(delegacion_id):
    delegacion = Delegacion.query.get_or_404(delegacion_id)

    # Delegado: solo su propia delegación
    if current_user.rol == 'delegado' and current_user.delegacion_id != delegacion.id:
        from flask import abort
        abort(403)

    if request.method == 'POST':
        # Solo admin/coordinador pueden crear CCTs
        if current_user.rol not in ('admin', 'coordinador'):
            from flask import abort
            abort(403)
        nuevo = Plantel(
            cct=request.form['cct'].strip().upper(),
            nombre=request.form.get('nombre').strip(),
            turno=request.form['turno'],
            nivel=request.form.get('nivel'),
            modalidad=request.form['modalidad'],
            zona_escolar=request.form['zona_escolar'],
            sector=request.form['sector'],
            calle=request.form.get('calle'),
            num_exterior=request.form.get('num_exterior'),
            num_interior=request.form.get('num_interior'),
            cruce_1=request.form.get('cruce_1'),
            cruce_2=request.form.get('cruce_2'),
            localidad=request.form.get('localidad'),
            colonia=request.form.get('colonia'),
            municipio=request.form.get('municipio'),
            cp=request.form.get('cp'),
            coordenadas_gps=request.form.get('coordenadas_gps'),
            estado='HIDALGO',
            delegacion_id=delegacion.id
        )
        db.session.add(nuevo)
        db.session.commit()
        flash('CCT registrado correctamente.', 'success')
        return redirect(url_for('delegaciones_bp.vista_ccts_por_delegacion', delegacion_id=delegacion.id))

    ccts = delegacion.planteles
    return render_template('consulta_ccts.html', delegacion=delegacion, ccts=ccts)


@delegaciones_bp.route('/eliminar_cct/<int:id>', methods=['POST'])
@roles_required('admin', 'coordinador')
def eliminar_cct(id):
    cct = Plantel.query.get_or_404(id)
    delegacion_id = cct.delegacion_id
    db.session.delete(cct)
    db.session.commit()
    registrar_notificacion(
        f"{current_user.nombre} eliminó el CCT {cct.cct} ({cct.nombre})",
        tipo="cct"
    )
    flash(f'CCT {cct.cct} eliminado.', 'warning')
    return redirect(url_for('delegaciones_bp.vista_ccts_por_delegacion', delegacion_id=delegacion_id))

@delegaciones_bp.route('/editar_cct/<int:id>', methods=['POST'])
@roles_required('admin', 'coordinador')
def editar_cct(id):
    cct = Plantel.query.get_or_404(id)
    cct.cct = request.form['nuevo_cct'].strip().upper()
    cct.nombre = request.form['nuevo_nombre'].strip()
    cct.turno = request.form['nuevo_turno'].strip()
    cct.nivel = request.form['nuevo_nivel'].strip()
    cct.modalidad = request.form['nuevo_modalidad'].strip()
    cct.zona_escolar = request.form['nuevo_zona_escolar'].strip()
    cct.sector = request.form['nuevo_sector'].strip()
    cct.calle = request.form.get('nuevo_calle', '').strip()
    cct.num_exterior = request.form.get('nuevo_num_exterior', '').strip()
    cct.num_interior = request.form.get('nuevo_num_interior', '').strip()
    cct.cruce_1 = request.form.get('nuevo_cruce_1', '').strip()
    cct.cruce_2 = request.form.get('nuevo_cruce_2', '').strip()
    cct.localidad = request.form.get('nuevo_localidad', '').strip()
    cct.colonia = request.form.get('nuevo_colonia', '').strip()
    cct.municipio = request.form.get('nuevo_municipio', '').strip()
    cct.cp = request.form.get('nuevo_cp', '').strip()
    cct.coordenadas_gps = request.form.get('nuevo_coordenadas_gps', '').strip()
    db.session.commit()
    flash('Plantel actualizado correctamente.', 'success')
    return redirect(url_for('delegaciones_bp.vista_ccts_por_delegacion', delegacion_id=cct.delegacion_id))

@delegaciones_bp.route('/descargar_plantilla_delegaciones')
@roles_required('admin')
def descargar_plantilla_delegaciones():
    return send_from_directory('static/plantillas', 'plantilla_delegaciones.xlsx', as_attachment=True)

@delegaciones_bp.route('/subir_excel/<int:delegacion_id>', methods=['POST'])
@roles_required('admin')
def subir_excel_delegaciones(delegacion_id):
    if 'archivo_excel' not in request.files:
        flash('No se envió ningún archivo.', 'danger')
        return redirect(url_for('delegaciones_bp.vista_delegaciones'))

    archivo = request.files['archivo_excel']

    if archivo.filename == '':
        flash('Nombre de archivo vacío.', 'danger')
        return redirect(url_for('delegaciones_bp.vista_delegaciones'))

    if archivo and archivo.filename.endswith('.xlsx'):
        try:
            df = pd.read_excel(archivo)
            registros_agregados = 0
            registros_ignorados = 0

            for index, row in df.iterrows():
                nombre = row.get('nombre')
                nivel = row.get('nivel')

                if isinstance(nombre, str) and isinstance(nivel, str) and nombre.strip() and nivel.strip():
                    nueva = Delegacion(nombre=nombre.strip(), nivel=nivel.strip().upper())
                    db.session.add(nueva)
                    registros_agregados += 1
                else:
                    registros_ignorados += 1

            db.session.commit()
            flash(f'Se cargaron {registros_agregados} delegaciones. {registros_ignorados} filas ignoradas.', 'success')
        except Exception as e:
            flash(f'Ocurrió un error al procesar el archivo: {str(e)}', 'danger')
    else:
        flash('Formato de archivo no permitido. Usa .xlsx', 'danger')

    return redirect(url_for('delegaciones_bp.vista_delegaciones'))

@delegaciones_bp.route('/delegacion/<int:delegacion_id>/subir_excel', methods=['POST'])
@roles_required('admin')
def subir_excel_ccts(delegacion_id):
    if 'archivo_excel' not in request.files:
        flash('No se envió ningún archivo.', 'danger')
        return redirect(url_for('delegaciones_bp.vista_ccts_por_delegacion', delegacion_id=delegacion_id))

    archivo = request.files['archivo_excel']

    if archivo.filename == '':
        flash('Nombre de archivo vacío.', 'danger')
        return redirect(url_for('delegaciones_bp.vista_ccts_por_delegacion', delegacion_id=delegacion_id))

    if archivo and archivo.filename.endswith('.xlsx'):
        try:
            df = pd.read_excel(archivo)
            registros_agregados = 0
            registros_ignorados = 0

            def limpiar(valor):
                return str(valor).strip() if pd.notna(valor) else ''

            for _, row in df.iterrows():
                cct = row.get('cct')
                nombre = row.get('nombre')
                turno = row.get('turno')
                nivel = row.get('nivel')
                modalidad = row.get('modalidad')

                if all(isinstance(v, str) and v.strip() for v in [cct, nombre, turno, nivel, modalidad]):
                    nuevo = Plantel(
                        cct=limpiar(cct).upper(),
                        nombre=limpiar(nombre),
                        turno=limpiar(turno),
                        nivel=limpiar(nivel),
                        modalidad=limpiar(modalidad),
                        zona_escolar=limpiar(row.get('zona_escolar')),
                        sector=limpiar(row.get('sector')),
                        calle=limpiar(row.get('calle')),
                        num_exterior=limpiar(row.get('num_exterior')),
                        num_interior=limpiar(row.get('num_interior')),
                        cruce_1=limpiar(row.get('cruce_1')),
                        cruce_2=limpiar(row.get('cruce_2')),
                        localidad=limpiar(row.get('localidad')),
                        colonia=limpiar(row.get('colonia')),
                        municipio=limpiar(row.get('municipio')),
                        cp=limpiar(row.get('cp')),
                        coordenadas_gps=limpiar(row.get('coordenadas_gps')),
                        estado='HIDALGO',
                        delegacion_id=delegacion_id
                    )
                    db.session.add(nuevo)
                    registros_agregados += 1
                else:
                    registros_ignorados += 1

            db.session.commit()
            flash(f'Se cargaron {registros_agregados} CCTs. {registros_ignorados} filas ignoradas.', 'success')
        except Exception as e:
            flash(f'Ocurrió un error al procesar el archivo: {str(e)}', 'danger')
    else:
        flash('Formato de archivo no permitido. Usa .xlsx', 'danger')

    return redirect(url_for('delegaciones_bp.vista_ccts_por_delegacion', delegacion_id=delegacion_id))

@delegaciones_bp.route("/delegaciones/reporte/excel")
@login_required
def reporte_delegaciones_excel():
    data = _fetch_delegaciones_data_para_reporte()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Encabezado
    ws.merge_cells("A1:E1")
    ws["A1"] = "REPORTE DE DELEGACIONES"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Generado por: {data['generado_por']}  |  Fecha: {data['generado_en']}"
    ws["A2"].alignment = center

    ws.append([])
    ws.append(["Total de delegaciones", data["total_delegaciones"]])

    # Totales por nivel
    ws.append(["Totales por nivel", "Cantidad"])
    ws["A5"].font = bold; ws["B5"].font = bold
    fila = 6
    for nivel, cant in sorted(data["totales_por_nivel"].items()):
        ws.cell(row=fila, column=1, value=nivel)
        ws.cell(row=fila, column=2, value=cant)
        fila += 1

    # Hoja detalle
    ws_det = wb.create_sheet("Detalle")
    headers = ["Delegación", "Nivel", "Delegado", "# CCTs"]
    ws_det.append(headers)
    for cell in ws_det[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for _, nombre, nivel, delegado, num_ccts in data["rows"]:
        ws_det.append([nombre, nivel, delegado if delegado else "—", num_ccts])

    # Bordes + anchos
    for row in ws_det.iter_rows(min_row=1, max_row=ws_det.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    for col_idx in range(1, 5):
        max_len = 0
        for row in range(1, ws_det.max_row + 1):
            val = ws_det.cell(row=row, column=col_idx).value
            max_len = max(max_len, len(str(val)) if val else 0)
        ws_det.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws_det.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"delegaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@delegaciones_bp.route("/delegaciones/reporte/pdf")
@login_required
def reporte_delegaciones_pdf():
    data = _fetch_delegaciones_data_para_reporte()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("<b>REPORTE DE DELEGACIONES</b>", styles["Title"]))
    meta = f"Generado por: {data['generado_por']} &nbsp;&nbsp;|&nbsp;&nbsp; Fecha: {data['generado_en']}"
    story.append(Paragraph(meta, styles["Normal"]))
    story.append(Spacer(1, 10))

    # Resumen
    story.append(Paragraph(f"<b>Total de delegaciones:</b> {data['total_delegaciones']}", styles["Heading3"]))

    # Totales por nivel
    resumen_tbl_data = [["Nivel", "Cantidad"]]
    for nivel, cant in sorted(data["totales_por_nivel"].items()):
        resumen_tbl_data.append([nivel, str(cant)])
    resumen_tbl = Table(resumen_tbl_data, hAlign="LEFT")
    resumen_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#ECECEC")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ]))
    story.append(resumen_tbl)
    story.append(Spacer(1, 14))

    # Detalle
    detail_data = [["Delegación", "Nivel", "Delegado", "# CCTs"]]
    for _, nombre, nivel, delegado, num_ccts in data["rows"]:
        detail_data.append([nombre, nivel, delegado if delegado else "—", str(num_ccts)])

    detail_tbl = Table(detail_data, repeatRows=1, hAlign="LEFT", colWidths=[220, 180, 220, 70])
    detail_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F5F5F5")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("ALIGN", (-1,1), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#FAFAFA")]),
    ]))
    story.append(detail_tbl)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    filename = f"delegaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(BytesIO(pdf), as_attachment=True, download_name=filename, mimetype="application/pdf")

@delegaciones_bp.route("/planteles/reporte/excel")
@login_required
def reporte_ccts_excel():
    data = _fetch_ccts_grouped_by_delegacion()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Encabezado
    ws.merge_cells("A1:D1")
    ws["A1"] = "REPORTE DE PLANTELES (CCTs)"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generado por: {data['generado_por']}  |  Fecha: {data['generado_en']}"
    ws["A2"].alignment = center

    ws.append([])
    ws.append(["Total de delegaciones", data["totales"]["delegaciones"]])
    ws.append(["Total de planteles", data["totales"]["planteles"]])
    ws.append([])
    ws.append(["Delegación", "Planteles"])
    ws["A6"].font = bold; ws["B6"].font = bold
    fila = 7
    for nombre, cant in data["totales"]["planteles_por_delegacion"].items():
        ws.cell(row=fila, column=1, value=nombre)
        ws.cell(row=fila, column=2, value=cant)
        fila += 1

    # Hoja por delegación
    headers = [
        "CCT", "Nombre", "Turno", "Nivel", "Modalidad",
        "Zona Escolar", "Sector",
        "Calle", "No. Ext", "No. Int",
        "Cruce 1", "Cruce 2",
        "Localidad", "Colonia", "Municipio", "CP",
        "Coordenadas GPS"
    ]

    for d in data["delegaciones"]:
        # Excel limita a 31 caracteres el nombre de hoja
        sheet_name = f"{d['nombre']}"[:31] if d['nombre'] else f"Deleg_{d['id']}"
        ws_det = wb.create_sheet(sheet_name)

        # Encabezado hoja
        ws_det.merge_cells("A1:R1")
        ws_det["A1"] = f"{d['nombre']} — {d['nivel']}"
        ws_det["A1"].font = Font(size=12, bold=True)
        ws_det["A1"].alignment = center

        delegado_txt = d["delegado"] if d["delegado"] else "—"
        ws_det.merge_cells("A2:R2")
        ws_det["A2"] = f"Delegado(a): {delegado_txt}    |    Planteles: {len(d['planteles'])}"
        ws_det["A2"].alignment = center

        ws_det.append([])

        ws_det.append(headers)
        for cell in ws_det[ws_det.max_row]:
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for p in d["planteles"]:
            ws_det.append([
                p["cct"], p["nombre"], p["turno"], p["nivel"], p["modalidad"],
                p["zona_escolar"], p["sector"],
                p["calle"], p["num_exterior"], p["num_interior"],
                p["cruce_1"], p["cruce_2"],
                p["localidad"], p["colonia"], p["municipio"], p["cp"],
                p["coordenadas_gps"]
            ])

        # Bordes y anchos
        for row in ws_det.iter_rows(min_row=4, max_row=ws_det.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border

        for col_idx in range(1, len(headers) + 1):
            max_len = 0
            for row_i in range(1, ws_det.max_row + 1):
                val = ws_det.cell(row=row_i, column=col_idx).value
                max_len = max(max_len, len(str(val)) if val else 0)
            ws_det.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)

        ws_det.freeze_panes = "A5"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"planteles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@delegaciones_bp.route("/planteles/reporte/pdf")
@login_required
def reporte_ccts_pdf():
    data = _fetch_ccts_grouped_by_delegacion()

    from io import BytesIO
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=20, rightMargin=20, topMargin=22, bottomMargin=22
    )
    styles = getSampleStyleSheet()

    # Estilos compactos
    small = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        wordWrap="CJK",      # <— permite corte “en cualquier lado”
        spaceAfter=0,
        spaceBefore=0,
    )
    small_bold = ParagraphStyle("small_bold", parent=small, fontName="Helvetica-Bold")

    h2 = ParagraphStyle(
        "h2",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        spaceAfter=6
    )

    def P(val, bold=False, maxlen=None):
        """Paragraph seguro: sin None, truncado opcional para PDF."""
        s = "" if val is None else str(val)
        if maxlen and len(s) > maxlen:
            s = s[:maxlen-1] + "…"
        return Paragraph(s, small_bold if bold else small)

    story = []

    # Portada / resumen
    story.append(Paragraph("<b>REPORTE DE PLANTELES (CCTs)</b>", styles["Title"]))
    meta = f"Generado por: {data['generado_por']} &nbsp;&nbsp;|&nbsp;&nbsp; Fecha: {data['generado_en']}"
    story.append(Paragraph(meta, styles["Normal"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"<b>Total de delegaciones:</b> {data['totales']['delegaciones']}", styles["Heading3"]))
    story.append(Paragraph(f"<b>Total de planteles:</b> {data['totales']['planteles']}", styles["Heading3"]))
    story.append(Spacer(1, 10))

    # Encabezados de ambas tablas
    headers_generales = [
        P("No.", True), P("CCT", True), P("Nombre", True), P("Turno", True),
        P("Nivel", True), P("Modalidad", True), P("Zona", True), P("Sector", True)
    ]
    headers_dom = [
        P("No.", True), P("Calle", True), P("No.Ext", True), P("No.Int", True),
        P("Cruce 1", True), P("Cruce 2", True),
        P("Localidad", True), P("Colonia", True), P("Municipio", True),
        P("CP", True), P("GPS", True)
    ]

    # Anchos compactos (cm)
    colw_generales = [0.8*cm, 2.0*cm, 6.0*cm, 1.5*cm, 2.3*cm, 2.5*cm, 1.5*cm, 1.5*cm]
    colw_dom = [0.8*cm, 4.0*cm, 1.2*cm, 1.2*cm, 2.3*cm, 2.3*cm, 2.3*cm, 2.3*cm, 3.0*cm, 1.5*cm, 3.0*cm]

    # Estilo de tabla común
    def base_style():
        return TableStyle([
            ("FONTSIZE", (0,0), (-1,-1), 7.5),
            ("LEADING", (0,0), (-1,-1), 9),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F5F5F5")),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#FAFAFA")]),
        ])

    for i, d in enumerate(data["delegaciones"]):
        delegado_txt = d["delegado"] if d["delegado"] else "—"
        title = f"{d['nombre']} — {d['nivel']}  (Delegado(a): {delegado_txt} | Planteles: {len(d['planteles'])})"
        story.append(Paragraph(title, h2))
        story.append(Spacer(1, 4))

        # -------- Tabla 1: Generales (mismas columnas SIEMPRE) --------
        tbl_gen = [headers_generales]
        for idx, p in enumerate(d["planteles"], start=1):
            row = [
                P(idx),                         # No.
                P(p.get("cct")),                # CCT
                P(p.get("nombre"), maxlen=120), # Nombre (limite prudente)
                P(p.get("turno")),
                P(p.get("nivel")),
                P(p.get("modalidad")),
                P(p.get("zona_escolar")),
                P(p.get("sector")),
            ]
            # seguridad: si por alguna razón cambia el largo, lo normalizamos:
            row = (row + [P("")]*8)[:8]
            tbl_gen.append(row)

        table_gen = Table(
            tbl_gen, repeatRows=1, hAlign="LEFT",
            colWidths=colw_generales, splitByRow=1   # <— clave para cortes limpios
        )
        ts_gen = base_style()
        ts_gen.add("ALIGN", (0,1), (0,-1), "CENTER")   # No.
        ts_gen.add("ALIGN", (1,1), (1,-1), "CENTER")   # CCT
        ts_gen.add("ALIGN", (3,1), (3,-1), "CENTER")   # Turno
        ts_gen.add("ALIGN", (6,1), (7,-1), "CENTER")   # Zona/Sector
        table_gen.setStyle(ts_gen)
        story.append(table_gen)
        story.append(Spacer(1, 6))

        # -------- Tabla 2: Domicilio (mismas columnas SIEMPRE) --------
        tbl_dom = [headers_dom]
        for idx, p in enumerate(d["planteles"], start=1):
            row = [
                P(idx),
                P(p.get("calle"), maxlen=120),
                P(p.get("num_exterior")), P(p.get("num_interior")),
                P(p.get("cruce_1")), P(p.get("cruce_2")),
                P(p.get("localidad")), P(p.get("colonia")),
                P(p.get("municipio")),
                P(p.get("cp")), P(p.get("coordenadas_gps"), maxlen=120),
            ]
            row = (row + [P("")]*11)[:11]
            tbl_dom.append(row)

        table_dom = Table(
            tbl_dom, repeatRows=1, hAlign="LEFT",
            colWidths=colw_dom, splitByRow=1       # <— evita desacomodos
        )
        ts_dom = base_style()
        ts_dom.add("ALIGN", (0,1), (0,-1), "CENTER")   # No.
        ts_dom.add("ALIGN", (2,1), (3,-1), "CENTER")   # No.Ext/No.Int
        ts_dom.add("ALIGN", (9,1), (9,-1), "CENTER")   # CP
        table_dom.setStyle(ts_dom)
        story.append(table_dom)

        if i < len(data["delegaciones"]) - 1:
            story.append(PageBreak())

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    filename = f"planteles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(BytesIO(pdf), as_attachment=True, download_name=filename, mimetype="application/pdf")

@delegaciones_bp.route("/personal/reporte/excel")
@login_required
def reporte_personal_excel():
    delegacion_id = request.args.get("delegacion_id", type=int)
    if current_user.rol == "delegado" and (not delegacion_id or delegacion_id != current_user.delegacion_id):
        abort(403)
    data = _fetch_personal_por_cct()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabecera
    ws.merge_cells("A1:E1")
    ws["A1"] = "REPORTE DE PERSONAL POR CCT"
    ws["A1"].font = Font(size=14, bold=True); ws["A1"].alignment = center
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Generado por: {data['generado_por']}  |  Fecha: {data['generado_en']}"
    ws["A2"].alignment = center
    ws.append([])

    fila = 4
    for d in data["delegaciones"]:
        ws.append([f"Delegación: {d['nombre']} — {d['nivel']}"])
        ws.append(["Hombres", "Mujeres", "Total"])
        for c in ws[fila+1]:
            c.font = bold; c.fill = header_fill; c.alignment = center; c.border = border
        ws.append([d["totales"]["hombres"], d["totales"]["mujeres"], d["totales"]["total"]])
        fila += 3
        ws.append([]); fila += 1

        # Hoja detalle por delegación (una por cada delegación)
        sheet_name = f"{d['nombre']}"[:31] if d['nombre'] else f"Deleg_{d['id']}"
        ws_det = wb.create_sheet(sheet_name)

        # Encabezado hoja
        titulo = f"{d['nombre']} — {d['nivel']}  |  H: {d['totales']['hombres']}  M: {d['totales']['mujeres']}  T: {d['totales']['total']}"
        ws_det.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6 + len(d["funciones_orden"]))
        ws_det.cell(row=1, column=1, value=titulo).font = Font(size=12, bold=True)
        ws_det.cell(row=1, column=1).alignment = center

        # Encabezados
        headers = ["CCT", "Plantel", "H", "M", "Total"]
        headers += d["funciones_orden"]  # columnas dinámicas por función
        ws_det.append(headers)
        for cell in ws_det[2]:
            cell.font = bold; cell.fill = header_fill; cell.alignment = center; cell.border = border

        # Filas
        for n in d["planteles"]:
            fila_vals = [n["cct"], n["plantel"], n["hombres"], n["mujeres"], n["total"]]
            for f in d["funciones_orden"]:
                fila_vals.append(n["funciones"].get(f, 0))
            ws_det.append(fila_vals)

        # Bordes y anchos
        col_count = 5 + len(d["funciones_orden"])
        for row in ws_det.iter_rows(min_row=2, max_row=ws_det.max_row, min_col=1, max_col=col_count):
            for c in row:
                c.border = border

        for col_idx in range(1, col_count + 1):
            # ancho automático aproximado
            max_len = 0
            for r in range(1, ws_det.max_row + 1):
                v = ws_det.cell(row=r, column=col_idx).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws_det.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        ws_det.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output); output.seek(0)
    filename = f"personal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@delegaciones_bp.route("/personal/reporte/pdf")
@login_required
def reporte_personal_pdf():
    delegacion_id = request.args.get("delegacion_id", type=int)
    if current_user.rol == "delegado" and (not delegacion_id or delegacion_id != current_user.delegacion_id):
        abort(403)
    data = _fetch_personal_por_cct()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=20, rightMargin=20, topMargin=22, bottomMargin=22)
    styles = getSampleStyleSheet()
    small = ParagraphStyle("small", parent=styles["Normal"], fontName="Helvetica", fontSize=7.5, leading=9, wordWrap="CJK")
    small_bold = ParagraphStyle("small_bold", parent=small, fontName="Helvetica-Bold")

    def P(txt, bold=False):
        return Paragraph("" if txt is None else str(txt), small_bold if bold else small)

    story = []
    story.append(Paragraph("<b>REPORTE DE PERSONAL POR CCT</b>", styles["Title"]))
    story.append(Paragraph(f"Generado por: {data['generado_por']} &nbsp;&nbsp;|&nbsp;&nbsp; Fecha: {data['generado_en']}", styles["Normal"]))
    story.append(Spacer(1, 10))

    for i, d in enumerate(data["delegaciones"]):
        story.append(Paragraph(f"<b>Delegación:</b> {d['nombre']} — {d['nivel']}", styles["Heading2"]))
        story.append(Paragraph(f"<b>Totales delegación</b> — H: {d['totales']['hombres']}  |  M: {d['totales']['mujeres']}  |  T: {d['totales']['total']}", styles["Heading3"]))
        story.append(Spacer(1, 6))

        # Tabla: CCT / Plantel / H / M / Total / Funciones dinámicas...
        headers = [P("No.", True), P("CCT", True), P("Plantel", True), P("H", True), P("M", True), P("T", True)]
        for f in d["funciones_orden"]:
            headers.append(P(f, True))

        tbl_data = [headers]
        for idx, n in enumerate(d["planteles"], start=1):
            row = [P(idx), P(n["cct"]), P(n["plantel"]), P(n["hombres"]), P(n["mujeres"]), P(n["total"])]
            for f in d["funciones_orden"]:
                row.append(P(n["funciones"].get(f, 0)))
            tbl_data.append(row)

        # colWidths compactos (ajusta si tienes muchas funciones)
        base_widths = [0.8*cm, 2.0*cm, 6.5*cm, 1.2*cm, 1.2*cm, 1.2*cm]
        func_cols = len(d["funciones_orden"])
        func_widths = [2.0*cm] * func_cols  # cada función a 2.0 cm
        col_widths = base_widths + func_widths

        table = Table(tbl_data, repeatRows=1, hAlign="LEFT", colWidths=col_widths, splitByRow=1)
        table.setStyle(TableStyle([
            ("FONTSIZE", (0,0), (-1,-1), 7.5),
            ("LEADING", (0,0), (-1,-1), 9),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F5F5F5")),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("ALIGN", (0,1), (0,-1), "CENTER"),  # No.
            ("ALIGN", (1,1), (1,-1), "CENTER"),  # CCT
            ("ALIGN", (3,1), (5,-1), "CENTER"),  # H M T
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#FAFAFA")]),
        ]))
        story.append(table)

        if i < len(data["delegaciones"]) - 1:
            story.append(PageBreak())

    doc.build(story)
    pdf = buffer.getvalue(); buffer.close()
    filename = f"personal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(BytesIO(pdf), as_attachment=True, download_name=filename, mimetype="application/pdf")

@delegaciones_bp.route('/agregar_delegacion', methods=['POST'])
@roles_required('admin', 'coordinador')
@login_required
def agregar_delegacion():
    nombre   = (request.form.get('nombre') or '').strip()
    nivel    = (request.form.get('nivel') or '').strip()   # libre
    delegado = (request.form.get('delegado') or '').strip()

    if not nombre or not nivel:
        flash('Nombre y nivel son obligatorios.', 'warning')
        return redirect(url_for('delegaciones_bp.vista_delegaciones'))

    try:
        nueva = Delegacion(nombre=nombre, nivel=nivel, delegado=delegado or None)
        db.session.add(nueva)
        db.session.commit()
        registrar_notificacion(
            f"{current_user.nombre} creó la delegación '{nombre}' (nivel {nivel})",
            tipo="delegacion"
        )
        flash('Delegación creada correctamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'No se pudo crear la delegación: {e}', 'danger')

    return redirect(url_for('delegaciones_bp.vista_delegaciones'))

@delegaciones_bp.route('/delegaciones/<int:delegacion_id>/tabla-personal')
@login_required
def tabla_personal_delegacion(delegacion_id):
    delegacion = Delegacion.query.get_or_404(delegacion_id)

    # 👇 Delegado: solo su propia delegación
    if current_user.rol == 'delegado' and current_user.delegacion_id != delegacion.id:
        from flask import abort
        abort(403)

    # permiso de ver: cualquiera con rol válido que ya uses
    can_edit = has_role('admin') or has_role('coordinador')
    return render_template(
        'delegaciones/tabla_personal.html',
        delegacion=delegacion,
        can_edit=can_edit
    )

# ---------- API: listar (GET remoto para Tabulator) ----------
@delegaciones_bp.route('/api/delegaciones/<int:delegacion_id>/personal')
@login_required
def api_listar_personal(delegacion_id):
    Delegacion.query.get_or_404(delegacion_id)
    if current_user.rol == "delegado" and current_user.delegacion_id != delegacion_id:
        abort(403)

    has_page = request.args.get("page") is not None
    has_size = request.args.get("size") is not None
    page, size, sorters, filters = _parse_tabulator_args(request) if (has_page or has_size) else (None, None, [], [])

    ALIAS = {
        "puesto": Personal.funcion_coordinacion,
        "estatus": Personal.estatus_membresia,
        "telefono": Personal.tel1,
        "correo": Personal.correo_electronico,
    }

    q = (Personal.query
         .join(Plantel, Personal.cct == Plantel.cct)
         .filter(Plantel.delegacion_id == delegacion_id))

    # filtros
    for f in filters:
        field = (f or {}).get("field"); value = (f or {}).get("value")
        if not field or value in (None, ""): continue
        col = getattr(Personal, field, None) or ALIAS.get(field)
        if col is not None: q = q.filter(col.ilike(f"%{value}%"))

    # orden
    for s in sorters:
        field = (s or {}).get("field"); direction = (s or {}).get("dir", "asc")
        col = getattr(Personal, field, None) or ALIAS.get(field)
        if col is not None: q = q.order_by(col.asc() if direction == "asc" else col.desc())

    # paginación
    if page is None or size is None:
        rows = q.all()
        total = len(rows)
        last_page = 1
    else:
        total = q.count()
        rows = q.offset((page - 1) * size).limit(size).all()
        last_page = max(1, ceil(total / size))

    cols = [c.name for c in Personal.__table__.columns]
    def to_dict(p): 
        d = {}
        for c in cols:
            v = getattr(p, c, None)
            d[c] = v.isoformat() if hasattr(v, "isoformat") else v
        return d

    return jsonify({"data": [to_dict(r) for r in rows], "total": total, "last_page": last_page})



@delegaciones_bp.route('/api/delegaciones/<int:delegacion_id>/personal/bulk-update', methods=['POST'])
@login_required
@roles_required('admin', 'coordinador')
def api_guardar_personal_bulk(delegacion_id):
    Delegacion.query.get_or_404(delegacion_id)

    payload = request.get_json(silent=True) or {}
    rows = payload.get('rows', [])
    if not isinstance(rows, list):
        abort(400, description="Formato inválido")

    debug = request.args.get("debug", type=int) == 1
    skips = []  # para depuración opcional

    # Campos que aceptará el guardado masivo (incluye alias usados en la UI)
    UPDATABLE = {
        # Identificación / base
        "cct", "apellido_paterno", "apellido_materno", "nombre", "genero",
        "curp", "rfc", "clave_presupuestal", "funcion_coordinacion", "funcion", "grado_estudios", "titulado",
        "fecha_ingreso", "fecha_baja_jubilacion", "estatus_membresia", "nombramiento",

        # Contacto / domicilio persona
        "domicilio", "numero", "localidad", "colonia", "municipio", "cp",
        "tel1", "tel2", "correo_electronico",

        # v2
        "num", "dp_num_int", "dp_cruce1", "dp_cruce2",

        # Datos de plantel “cacheados” en personal (si los usas en UI)
        "escuela_nombre", "turno", "nivel", "subs_modalidad", "zona_escolar", "sector",
        "dom_esc_calle", "dom_esc_num_ext", "dom_esc_num_int", "dom_esc_cruce1", "dom_esc_cruce2",
        "dom_esc_localidad", "dom_esc_colonia", "dom_esc_mun_nom", "dom_esc_cp", "dom_esc_coordenadas_gps",

        # Otros
        "estado", "seccion_snte", "del_o_ct", "org", "coord_reg", "fun_sin",

        # Alias expuestos en la tabla
        "puesto", "estatus", "telefono", "correo"
    }

    # Alias de la UI → atributo real del modelo
    SETATTR_ALIAS = {
        "puesto": "funcion_coordinacion",     # 👈 el alias 'puesto' edita funcion_coordinacion
        "estatus": "estatus_membresia",
        "telefono": "tel1",
        "correo": "correo_electronico",
    }

    # helper: normaliza "" -> None (opcional, ajusta si quieres guardar "")
    def _norm_empty(val):
        if isinstance(val, str):
            v = val.strip()
            return v if v != "" else None
        return val

    # helper: compara updated_at tolerando tz/precisión
    def _dt_sig(s: str):
        # toma solo 'YYYY-mm-ddTHH:MM:SS'
        if not s or not isinstance(s, str):
            return None
        return s[:19]

    # ✅ catálogo permitido para funcion_coordinacion + normalización
    ALLOWED_FUNC_COORD = {
        "DIRECTOR (A)",
        "DOCENTE",
        "ADMINISTRATIVO (A)",
        "PREFECTO (A)",
        "INTENDENCIA",
    }
    def _canon_func_coord(val):
        """Normaliza variantes y valida contra el set permitido."""
        if val is None:
            return None
        s = str(val).strip().upper()

        # Normaliza paréntesis y espacios
        s = s.replace("DIRECTOR(A)", "DIRECTOR (A)")
        s = s.replace("ADMINISTRATIVO(A)", "ADMINISTRATIVO (A)")
        s = s.replace("PREFECTO(A)", "PREFECTO (A)")
        s = " ".join(s.split())

        # Sinónimos frecuentes → forma canónica
        synonyms = {
            "DIRECTOR": "DIRECTOR (A)",
            "DIRECTORA": "DIRECTOR (A)",
            "PROFESOR": "DOCENTE",
            "MAESTRO": "DOCENTE",
            "MAESTRA": "DOCENTE",
            "ADMINISTRATIVO": "ADMINISTRATIVO (A)",
            "PREFECTO": "PREFECTO (A)",
            "PREFECTA": "PREFECTO (A)",
            "INTENDENTE": "INTENDENCIA",
        }
        s = synonyms.get(s, s)

        return s if s in ALLOWED_FUNC_COORD else None

    actualizados = 0

    for r in rows:
        # --- ID robusto ---
        pid_raw = r.get("id")
        try:
            pid = int(pid_raw)
        except (TypeError, ValueError):
            if debug: skips.append({"id": pid_raw, "razon": "id_invalido"})
            continue

        # Asegura que la persona pertenezca a la delegación (JOIN con Plantel)
        persona = (Personal.query
                   .join(Plantel, Personal.cct == Plantel.cct)
                   .filter(Personal.id == pid, Plantel.delegacion_id == delegacion_id)
                   .first())
        if not persona:
            if debug: skips.append({"id": pid, "razon": "no_pertenece_delegacion_o_no_existe"})
            continue

        # Bloqueo optimista (tolerante)
        client_updated_at = r.get("updated_at")
        server_updated_at = getattr(persona, "updated_at", None)
        if client_updated_at and server_updated_at:
            if _dt_sig(server_updated_at.isoformat()) != _dt_sig(client_updated_at):
                if debug: skips.append({"id": pid, "razon": "conflicto_updated_at"})
                continue

        cambios = []
        for k, v in r.items():
            if k not in UPDATABLE:
                continue

            attr = SETATTR_ALIAS.get(k, k)  # alias → real
            if not hasattr(persona, attr):
                if debug: skips.append({"id": pid, "campo": k, "razon": "attr_no_existe"})
                continue

            # Normalizaciones ligeras
            v = _norm_empty(v)

            if attr in ("curp", "rfc") and isinstance(v, str):
                v = v.strip().upper() or None

            if attr in ("fecha_ingreso", "fecha_baja_jubilacion") and isinstance(v, str) and v:
                from datetime import datetime as _dt
                try:
                    v = _dt.strptime(v[:10], "%Y-%m-%d").date()
                except ValueError:
                    if debug: skips.append({"id": pid, "campo": attr, "razon": "fecha_invalida"})
                    continue

            # Validación FK del CCT si cambia
            if attr == "cct" and v and v != (persona.cct or ""):
                if not Plantel.query.filter_by(cct=v).first():
                    if debug: skips.append({"id": pid, "campo": "cct", "razon": "cct_inexistente"})
                    continue

            # ✅ Validación catálogo funcion_coordinacion
            if attr == "funcion_coordinacion":
                if v is not None:
                    canon = _canon_func_coord(v)
                    if not canon:
                        if debug: skips.append({"id": pid, "campo": attr, "razon": "valor_no_permitido"})
                        continue
                    v = canon

            prev = getattr(persona, attr, None)
            if v != prev:
                setattr(persona, attr, v)
                cambios.append((attr, prev, v))

        if cambios:
            if hasattr(persona, "updated_at"):
                persona.updated_at = datetime.now(timezone('America/Mexico_City'))
            db.session.add(persona)
            db.session.flush()

            for campo, antes, despues in cambios:
                try:
                    registrar_historial(
                        entidad="personal",
                        campo=campo,
                        valor_anterior=antes,
                        valor_nuevo=despues,
                        entidad_id=persona.id,
                        usuario=getattr(current_user, "nombre", "sistema"),
                        tipo="edicion masiva"
                    )
                except Exception:
                    pass
            actualizados += 1
        else:
            if debug: skips.append({"id": pid, "razon": "sin_cambios"})

    db.session.commit()
    resp = {"ok": True, "actualizados": actualizados}
    if debug:
        resp["skips"] = skips
    return jsonify(resp)



@delegaciones_bp.route('/api/delegaciones/<int:delegacion_id>/personal/summary')
@login_required
def api_resumen_personal(delegacion_id):
    Delegacion.query.get_or_404(delegacion_id)
    if current_user.rol == "delegado" and current_user.delegacion_id != delegacion_id:
        from flask import abort
        abort(403)

    # 👉 lee parámetro (por si un día quieres incluirlos desde la UI)
    excluir_en_proceso = (request.args.get("excluir_baja_en_proceso", "1") == "1")

    q = (
        db.session.query(Personal.funcion_coordinacion, Personal.genero, func.count(Personal.id))
        .join(Plantel, Personal.cct == Plantel.cct)
        .filter(Plantel.delegacion_id == delegacion_id)
        .filter(not_(Personal.estatus_membresia.in_(['BAJA EN PROCESO','BAJA'])))
        .group_by(Personal.funcion_coordinacion, Personal.genero)
    )


    if excluir_en_proceso:
        q = q.filter(Personal.estatus_membresia.notin_(['BAJA EN PROCESO','BAJA']))

    # (opcional) si también quieres excluir BAJA definitiva:
    # q = q.filter(Personal.estatus_membresia.notin_(['BAJA EN PROCESO','BAJA']))

    q = q.group_by(Personal.funcion_coordinacion, Personal.genero).all()

    funciones_map = {}
    tot_h = tot_m = 0
    for funcion_coord, genero, cnt in q:
        f = (funcion_coord or "SIN FUNCIÓN COORD.").upper()
        g = (genero or "").upper()
        nodo = funciones_map.setdefault(f, {"funcion": f, "hombres": 0, "mujeres": 0, "total": 0})
        cnt = int(cnt)
        if g == "H":
            nodo["hombres"] += cnt; tot_h += cnt
        elif g == "M":
            nodo["mujeres"] += cnt; tot_m += cnt
        nodo["total"] += cnt

    tot_total = sum(v["total"] for v in funciones_map.values())
    funciones_list = sorted(funciones_map.values(), key=lambda x: x["funcion"])

    return jsonify({
        "totales": {"hombres": tot_h, "mujeres": tot_m, "total": tot_total},
        "funciones": funciones_list
    })


@delegaciones_bp.route('/api/delegaciones/<int:delegacion_id>/personal/export-excel', endpoint='api_exportar_personal_excel')
@login_required
def api_exportar_personal_excel(delegacion_id):
    # Validación de alcance
    Delegacion.query.get_or_404(delegacion_id)
    if current_user.rol == "delegado" and current_user.delegacion_id != delegacion_id:
        from flask import abort
        abort(403)

    # Base query: solo personal de la delegación
    base_query = (Personal.query
                  .join(Plantel, Personal.cct == Plantel.cct)
                  .filter(Plantel.delegacion_id == delegacion_id))

    # ⛔ EXCLUIR BAJA EN PROCESO y BAJA (hazlo aquí; NO uses 'query' todavía)
    base_query = base_query.filter(~Personal.estatus_membresia.in_(["BAJA EN PROCESO", "BAJA"]))

    # Alias opcionales (compatibles con la UI)
    ALIAS = {
        "puesto": Personal.funcion_coordinacion,
        "estatus": Personal.estatus_membresia,
        "telefono": Personal.tel1,
        "correo": Personal.correo_electronico,
    }

    # Reusar sorters/filters que manda Tabulator
    _, _, sorters, filters = _parse_tabulator_args(request)

    # --- Aplicar filtros (para datos y para resumen) ---
    def apply_filters(q):
        for f in filters:
            field = (f or {}).get("field")
            value = (f or {}).get("value")
            if not field or value in (None, ""):
                continue
            col = getattr(Personal, field, None) or ALIAS.get(field)
            if col is None:
                continue
            q = q.filter(col.ilike(f"%{value}%"))
        return q

    # ==== Hoja 1: Datos ====
    query = apply_filters(base_query)

    # Orden
    for s in sorters:
        field = (s or {}).get("field")
        direction = (s or {}).get("dir", "asc")
        col = getattr(Personal, field, None) or ALIAS.get(field)
        if col is None:
            continue
        query = query.order_by(col.asc() if direction == "asc" else col.desc())

    rows = query.all()

    # ---------- Columnas a exportar ----------
    ALL_COLS = [c.name for c in Personal.__table__.columns]
    HIDE_FIELDS = {"num", "id"}

    DISPLAY_ORDER = [
        "apellido_paterno","apellido_materno","nombre","genero","rfc","curp",
        "clave_presupuestal","funcion_coordinacion","funcion","grado_estudios","titulado",
        "fecha_ingreso","fecha_baja_jubilacion","estatus_membresia","nombramiento",
        "domicilio","numero","dp_num_int","dp_cruce1","dp_cruce2",
        "localidad","colonia","municipio","cp","tel1","tel2","correo_electronico",
        "escuela_nombre","cct","turno","nivel","subs_modalidad","zona_escolar","sector",
        "dom_esc_calle","dom_esc_num_ext","dom_esc_num_int","dom_esc_cruce1","dom_esc_cruce2",
        "dom_esc_localidad","dom_esc_colonia","dom_esc_mun_nom","dom_esc_cp","dom_esc_coordenadas_gps",
        "estado","seccion_snte","del_o_ct","org","coord_reg","fun_sin",
        "updated_at"
    ]

    ordered_cols = [c for c in DISPLAY_ORDER if c in ALL_COLS and c not in HIDE_FIELDS]
    for c in ALL_COLS:
        if c not in ordered_cols and c not in HIDE_FIELDS:
            ordered_cols.append(c)

    EXCEL_TITLES = {
        "apellido_paterno":"PATERNO","apellido_materno":"MATERNO","nombre":"NOMBRE",
        "genero":"GENERO","rfc":"RFC","curp":"CURP","clave_presupuestal":"CLAVE_PRESUPUESTAL",
        "funcion_coordinacion":"FUNCION_COORDINACION","funcion":"FUNCION","grado_estudios":"GRADO_MAXIMO_ESTUDIOS","titulado":"TITULADO",
        "fecha_ingreso":"FECHA_INGRESO","fecha_baja_jubilacion":"FCH_BAJ_JUB",
        "estatus_membresia":"STATUS_MEMB","nombramiento":"NOMBRAMIENTO",
        "domicilio":"DP_CALLE","numero":"DP_NUM_EXT","dp_num_int":"DP_NUM_INT",
        "dp_cruce1":"DP_CRUCE1","dp_cruce2":"DP_CRUCE2","localidad":"DP_LOCALIDAD",
        "colonia":"DP_COLONIA","municipio":"DP_MUN_NOM","cp":"DP_CP",
        "tel1":"DP_TEL1","tel2":"DP_TEL2","correo_electronico":"CORREO_ELECTRONICO",
        "escuela_nombre":"ESCUELA_NOMBRE","cct":"CCT","turno":"TURNO","nivel":"NIVEL",
        "subs_modalidad":"SUBS_MODALIDAD","zona_escolar":"ZONA_ESCOLAR","sector":"SECTOR",
        "dom_esc_calle":"DOM_ESC_CALLE","dom_esc_num_ext":"DOM_ESC_NUM_EXT","dom_esc_num_int":"DOM_ESC_NUM_INT",
        "dom_esc_cruce1":"DOM_ESC_CRUCE1","dom_esc_cruce2":"DOM_ESC_CRUCE2",
        "dom_esc_localidad":"DOM_ESC_LOCALIDAD","dom_esc_colonia":"DOM_ESC_COLONIA",
        "dom_esc_mun_nom":"DOM_ESC_MUN_NOM","dom_esc_cp":"DOM_ESC_CP","dom_esc_coordenadas_gps":"DOM_ESC_COORDENADAS GPS",
        "estado":"ESTADO","seccion_snte":"SECCION_SNTE","del_o_ct":"DEL_O_CT","org":"ORG","coord_reg":"COORD_REG","fun_sin":"FUN_SIN",
        "updated_at":"UPDATED_AT"
    }

    # ---------- Construir Excel ----------
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import date, datetime as dt

    wb = Workbook()

    # Hoja 1: Datos
    ws = wb.active
    ws.title = "Personal"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append([EXCEL_TITLES.get(c, c.upper()) for c in ordered_cols])
    for cell in ws[1]:
        cell.font = bold; cell.alignment = center; cell.fill = header_fill; cell.border = border

    for p in rows:
        vals = []
        for c in ordered_cols:
            v = getattr(p, c, "")
            if isinstance(v, (dt, date)):
                v = v.isoformat()
            vals.append(v)
        ws.append(vals)

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(ordered_cols)):
        for cell in r:
            cell.border = border

    for idx, c in enumerate(ordered_cols, start=1):
        max_len = len(str(EXCEL_TITLES.get(c, c)))
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=idx).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"

    # ==== Hoja 2: Resumen ====
    ws2 = wb.create_sheet("Resumen")

    sum_query = (db.session.query(Personal.funcion_coordinacion, Personal.genero, func.count(Personal.id))
                 .join(Plantel, Personal.cct == Plantel.cct)
                 .filter(Plantel.delegacion_id == delegacion_id))

    # 💡 Mismo filtro de exclusión en el resumen
    sum_query = sum_query.filter(~Personal.estatus_membresia.in_(["BAJA EN PROCESO", "BAJA"]))

    sum_query = apply_filters(sum_query)

    funciones_map, tot_h, tot_m = {}, 0, 0
    for funcion_coord, genero, cnt in sum_query.group_by(Personal.funcion_coordinacion, Personal.genero).all():
        f = (funcion_coord or "SIN FUNCIÓN COORD.").upper()
        g = (genero or "").upper()
        nodo = funciones_map.setdefault(f, {"hombres": 0, "mujeres": 0, "total": 0})
        cnt = int(cnt)
        if g == "H":
            nodo["hombres"] += cnt; tot_h += cnt
        elif g == "M":
            nodo["mujeres"] += cnt; tot_m += cnt
        nodo["total"] += cnt

    tot_total = sum(v["total"] for v in funciones_map.values())

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "RESUMEN POR FUNCIÓN DE COORDINACIÓN"
    ws2["A1"].font = Font(size=14, bold=True)
    ws2["A1"].alignment = center

    ws2.append([])
    ws2.append(["Hombres", "Mujeres", "Total"])
    for c in ws2[3]:
        c.font = bold; c.alignment = center; c.fill = header_fill; c.border = border
    ws2.append([tot_h, tot_m, tot_total])

    ws2.append([])
    ws2.append(["Función de coordinación", "Hombres", "Mujeres", "Total"])
    for c in ws2[6]:
        c.font = bold; c.alignment = center; c.fill = header_fill; c.border = border

    row_i = 7
    for funcion in sorted(funciones_map.keys()):
        vals = funciones_map[funcion]
        ws2.append([funcion, vals["hombres"], vals["mujeres"], vals["total"]])
        for c in ws2[row_i]:
            c.border = border
        row_i += 1

    for col_idx in range(1, 5):
        max_len = 0
        for r in range(1, ws2.max_row + 1):
            val = ws2.cell(row=r, column=col_idx).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    output = BytesIO()
    wb.save(output); output.seek(0)
    filename = f"personal_deleg_{delegacion_id}_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
